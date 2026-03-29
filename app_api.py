import copy
import importlib
import json
import os
import re
import sys
import threading
import time
import uuid
from pathlib import Path

from email_channel import resolve_channel
from frontend_run_context import ensure_run_context_dirs, load_run_context, serialize_run_context
from user_settings import (
    UserSettingsStore,
    ensure_directory,
    get_default_save_path as resolve_default_save_path,
    get_default_debug_trace_path,
    get_output_state_dir,
    get_packaged_diagnostics_dir,
)


class QuotaExceededError(RuntimeError):
    pass


class _FallbackDocumentTraceStore:
    """In-memory trace store used when the optional diagnostics module is absent."""

    def __init__(self, output_path=None):
        self.output_path = os.path.abspath(
            output_path or get_default_debug_trace_path()
        )
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self._records = {}
        self._order = []
        self._archive_index = {}

    def start_document(self, source_filename, source_path=None, document_id=None):
        doc_id = document_id or uuid.uuid4().hex
        if doc_id not in self._records:
            self._records[doc_id] = {
                "source_filename": source_filename,
                "document_id": doc_id,
                "extractor_raw_result": None,
                "normalized_fields": None,
                "classification_result": None,
                "naming_result": None,
                "combine_keys": None,
                "combine_result": None,
                "archive_target": None,
                "failure_reason": None,
            }
            if source_path:
                self._records[doc_id]["source_path"] = source_path
            self._order.append(doc_id)
        return doc_id

    def get_record(self, document_id):
        return self._records.get(document_id)

    def iter_records(self):
        for document_id in self._order:
            yield self._records[document_id]

    def set_fields(self, document_id, **fields):
        record = self._records.get(document_id)
        if not record:
            return

        for key, value in fields.items():
            record[key] = copy.deepcopy(value)

        archive_target = fields.get("archive_target")
        if archive_target:
            self.bind_archive_target(document_id, archive_target)

    def bind_archive_target(self, document_id, archive_target):
        record = self._records.get(document_id)
        if not record or not archive_target:
            return

        old_target = record.get("archive_target")
        if old_target:
            self._archive_index.pop(self._normalize_path(old_target), None)

        normalized = self._normalize_path(archive_target)
        self._archive_index[normalized] = document_id
        record["archive_target"] = archive_target

    def move_archive_target(self, old_target, new_target):
        if not old_target or not new_target:
            return None

        normalized_old = self._normalize_path(old_target)
        document_id = self._archive_index.pop(normalized_old, None)
        if not document_id:
            return None

        normalized_new = self._normalize_path(new_target)
        self._archive_index[normalized_new] = document_id
        self._records[document_id]["archive_target"] = new_target
        return document_id

    def get_document_id_by_archive_target(self, archive_target):
        if not archive_target:
            return None
        return self._archive_index.get(self._normalize_path(archive_target))

    def record_failure_event(self, document_id, code, stage, message=None, severity="failure"):
        record = self._records.get(document_id)
        if not record:
            return

        event = {
            "code": code,
            "stage": stage,
            "severity": severity,
        }
        if message:
            event["message"] = message

        failure_reason = record.get("failure_reason")
        if not failure_reason:
            record["failure_reason"] = dict(event)
            record["failure_reason"]["history"] = [dict(event)]
            return

        history = failure_reason.setdefault("history", [])
        history.append(dict(event))

        if self._severity_rank(severity) >= self._severity_rank(failure_reason.get("severity")):
            failure_reason.update(event)

    def flush(self):
        temp_path = f"{self.output_path}.tmp"
        with open(temp_path, "w", encoding="utf-8") as handle:
            for document_id in self._order:
                json.dump(
                    self._records[document_id],
                    handle,
                    ensure_ascii=False,
                    default=str,
                )
                handle.write("\n")
        os.replace(temp_path, self.output_path)

    @staticmethod
    def _normalize_path(path):
        return os.path.normcase(os.path.abspath(path))

    @staticmethod
    def _severity_rank(severity):
        ranks = {
            None: 0,
            "info": 1,
            "skipped": 2,
            "fallback": 3,
            "failure": 4,
        }
        return ranks.get(severity, 0)

class InvoiceAppAPI:
    def __init__(self):
        self._run_context = load_run_context()
        ensure_run_context_dirs(self._run_context)
        self._diag_lock = threading.Lock()
        self._packaged_diag_enabled = bool(getattr(sys, "frozen", False))
        self._packaged_diag_poll_count = 0
        self._packaged_diag_last_progress_signature = None
        self._packaged_diag_excepthook_installed = False
        self.progress = 0
        self.status_text = "Ready to start..."
        self.logs = []
        self._is_running = False
        self._stop_requested = False
        self.run_state = "idle"
        self.last_error = ""
        self.quota_exhausted = False
        self.quota_message = ""
        self._worker_thread = None
        self._settings_store = UserSettingsStore()
        self._requested_save_path = ""
        self._effective_save_path = ""
        self._effective_date_from = ""
        self._effective_date_to = ""
        self._current_run_id = self._run_context.get("run_id", "")
        self.audit_counts = {"manual_check": 0, "retention": 0, "raw_invoices": 0}
        self.discovered_categories = set()
        self.processed_invoices = []
        self.error_invoices = []
        self.stats = {"emails": 0, "invoices": 0, "errors": 0}
        self._email_level_url_evidence_seen = set()
        self._last_export_path = ""
        self._install_packaged_thread_excepthook()
        # Release-prep candidate: startup must not trigger local developer scans.

    def _refresh_run_context(self):
        refreshed_context = load_run_context()
        if refreshed_context != self._run_context:
            self._run_context = refreshed_context
        ensure_run_context_dirs(self._run_context)
        self._current_run_id = self._run_context.get("run_id", "")
        return self._run_context

    def _monitoring_path(self, filename):
        run_context = self._refresh_run_context()
        monitoring_dir = run_context.get("monitoring_dir", "")
        if not monitoring_dir:
            return ""
        return os.path.join(monitoring_dir, filename)

    def _snapshot_counts(self):
        return {
            "emails": int(getattr(self, "stats", {}).get("emails", 0) or 0),
            "archived": int(getattr(self, "stats", {}).get("invoices", 0) or 0),
            "manual_check": int(getattr(self, "audit_counts", {}).get("manual_check", 0) or 0),
            "retention": int(getattr(self, "audit_counts", {}).get("retention", 0) or 0),
            "raw_invoices": int(getattr(self, "audit_counts", {}).get("raw_invoices", 0) or 0),
            "errors": int(getattr(self, "stats", {}).get("errors", 0) or 0),
        }

    def _diag_append_jsonl(self, path, payload):
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with self._diag_lock:
                with open(path, "a", encoding="utf-8") as fh:
                    fh.write(json.dumps(payload, ensure_ascii=False, default=str))
                    fh.write("\n")
        except Exception as exc:
            print(f"[diag] append jsonl failed: {path}: {exc}")

    def _diag_write_json(self, path, payload):
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            temp_path = f"{path}.tmp"
            with self._diag_lock:
                with open(temp_path, "w", encoding="utf-8") as fh:
                    json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
                os.replace(temp_path, path)
        except Exception as exc:
            print(f"[diag] write json failed: {path}: {exc}")

    def _packaged_diag_dir(self):
        if not self._packaged_diag_enabled:
            return ""
        try:
            return get_packaged_diagnostics_dir()
        except Exception:
            return ""

    def _packaged_diag_file(self):
        diag_dir = self._packaged_diag_dir()
        if not diag_dir:
            return ""
        return os.path.join(diag_dir, "packaged_5p_diag.jsonl")

    def _packaged_diag_email_domain(self, email_address):
        email_text = str(email_address or "").strip()
        if "@" not in email_text:
            return ""
        return email_text.split("@", 1)[1].lower()

    def _sensitive_summary(self, email_address="", auth_code="", api_key=""):
        return {
            "email_domain": self._packaged_diag_email_domain(email_address),
            "has_auth_code": bool(auth_code),
            "has_api_key": bool(api_key),
        }

    def _create_document_trace_store(self, output_path=None):
        try:
            trace_module = importlib.import_module("diagnostics_trace")
            trace_store_cls = getattr(trace_module, "DocumentTraceStore")
            return trace_store_cls(output_path=output_path)
        except Exception as exc:
            self._packaged_diag_write(
                "document_trace_store_fallback",
                "_run_processing_loop",
                "exception",
                summary={"include_traceback": False},
                exc=exc,
            )
            return _FallbackDocumentTraceStore(output_path=output_path)

    def _packaged_diag_summary(self, summary):
        if not isinstance(summary, dict):
            return {}

        allowed_keys = {
            "requested_save_path",
            "effective_save_path",
            "date_from",
            "date_to",
            "email_domain",
            "has_auth_code",
            "has_api_key",
            "email_count",
            "attachment_count",
            "file_name",
            "thread_is_alive",
            "target_progress",
            "attachments",
            "poll_index",
            "logs_count",
            "include_traceback",
            "thread_target",
            "thread_target_ident",
            "connect_result",
            "scan_result_count",
            "extract_result_count",
            "url_result_count",
        }
        return {key: value for key, value in summary.items() if key in allowed_keys}

    def _packaged_diag_write(self, stage, function_name, outcome, summary=None, exc=None):
        if not self._packaged_diag_enabled:
            return

        exc_type = ""
        exc_message = ""
        if exc is not None:
            exc_type = getattr(getattr(exc, "__class__", None), "__name__", "") or type(exc).__name__
            exc_message = str(exc)

        self._diag_append_jsonl(
            self._packaged_diag_file(),
            {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                "stage": stage,
                "function": function_name,
                "outcome": outcome,
                "thread_name": threading.current_thread().name,
                "thread_ident": threading.current_thread().ident,
                "pid": os.getpid(),
                "run_state": self.run_state,
                "progress": self.progress,
                "status_text": self.status_text,
                "summary": self._packaged_diag_summary(summary),
                "exc_type": exc_type,
                "exc_message": exc_message,
            },
        )

    def _packaged_diag_reset(self, summary=None):
        if not self._packaged_diag_enabled:
            return

        diag_file = self._packaged_diag_file()
        if not diag_file:
            return

        try:
            os.makedirs(os.path.dirname(diag_file), exist_ok=True)
            with self._diag_lock:
                with open(diag_file, "w", encoding="utf-8"):
                    pass
        except Exception as exc:
            print(f"[diag] reset packaged jsonl failed: {diag_file}: {exc}")
            return

        self._packaged_diag_poll_count = 0
        self._packaged_diag_last_progress_signature = None
        self._packaged_diag_write("session_start", "start_processing", "success", summary=summary)

    def _install_packaged_thread_excepthook(self):
        if not self._packaged_diag_enabled or self._packaged_diag_excepthook_installed:
            return

        previous_hook = getattr(threading, "excepthook", None)

        def _hook(args):
            try:
                thread_obj = getattr(args, "thread", None)
                self._packaged_diag_write(
                    "thread_unhandled_exception",
                    "threading.excepthook",
                    "exception",
                    summary={
                        "thread_target": getattr(thread_obj, "name", ""),
                        "thread_target_ident": getattr(thread_obj, "ident", None),
                    },
                    exc=getattr(args, "exc_value", None),
                )
            except Exception:
                pass

            if previous_hook:
                previous_hook(args)

        threading.excepthook = _hook
        self._packaged_diag_excepthook_installed = True

    def _packaged_diag_log_progress_poll(self, payload):
        if not self._packaged_diag_enabled:
            return

        self._packaged_diag_poll_count += 1
        signature = (
            payload.get("progress"),
            payload.get("run_state"),
            payload.get("status_text"),
            payload.get("last_error"),
        )
        should_log = self._packaged_diag_poll_count <= 20 or signature != self._packaged_diag_last_progress_signature
        self._packaged_diag_last_progress_signature = signature
        if not should_log:
            return

        self._packaged_diag_write(
            "progress_poll",
            "get_progress",
            "success",
            summary={
                "poll_index": self._packaged_diag_poll_count,
                "logs_count": len(payload.get("logs", []) or []),
            },
        )

    def _safe_emit_run_state_event(self, from_state, to_state):
        if not self._run_context.get("enabled"):
            return
        self._diag_append_jsonl(
            self._monitoring_path("run_state_events.jsonl"),
            {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                "from_state": from_state,
                "to_state": to_state,
                "progress": self.progress,
                "status_text": self.status_text,
                "last_error": self.last_error,
                **self._snapshot_counts(),
            },
        )

    def _safe_emit_stage_event(self, stage, event, extra=None):
        if not self._run_context.get("enabled"):
            return
        payload = {
            "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
            "stage": stage,
            "event": event,
            "run_state": self.run_state,
            "progress": self.progress,
            "status_text": self.status_text,
            **self._snapshot_counts(),
        }
        if extra:
            payload.update(extra)
        self._diag_append_jsonl(self._monitoring_path("stage_events.jsonl"), payload)

    def _safe_emit_artifact_event(self, kind, path, document_id=None, source_kind=None, reason_code=None, category=None, extra=None):
        if not self._run_context.get("enabled"):
            return
        payload = {
            "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
            "kind": kind,
            "path": path,
            "document_id": document_id,
            "source_kind": source_kind,
            "reason_code": reason_code,
            "category": category,
            **self._snapshot_counts(),
        }
        if extra:
            payload.update(extra)
        self._diag_append_jsonl(self._monitoring_path("artifact_events.jsonl"), payload)

    def _safe_emit_input_inventory_event(self, payload):
        if not self._run_context.get("enabled"):
            return
        self._diag_append_jsonl(
            self._monitoring_path("input_attachment_inventory.jsonl"),
            {
                "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                **payload,
            },
        )

    def _attachment_diag_metadata(self, info, file_name=None, document_id=None, extra=None):
        payload = {
            "document_id": document_id,
            "email_id": info.get("email_id"),
            "sender": info.get("sender"),
            "subject": info.get("subject", ""),
            "file_name": file_name or info.get("file_name") or os.path.basename(str(info.get("filepath", ""))),
            "original_filename": info.get("original_filename"),
            "attachment_ext": info.get("attachment_ext"),
            "payload_size": info.get("payload_size"),
            "mime_content_type": info.get("mime_content_type"),
            "content_disposition": info.get("content_disposition"),
            "attachment_pair_key": info.get("attachment_pair_key"),
            "sibling_pdf_present": info.get("sibling_pdf_present"),
            "sibling_ofd_present": info.get("sibling_ofd_present"),
            "sibling_xml_present": info.get("sibling_xml_present"),
            "provider_unzipped_pair_suspected": info.get("provider_unzipped_pair_suspected"),
            "zip_context": info.get("zip_context"),
            "candidate_bucket": info.get("candidate_bucket"),
            "candidate_action": info.get("candidate_action"),
            "source_kind": info.get("source_kind"),
            "prefilter_reason_code": info.get("prefilter_reason_code"),
            "prefilter_signals": info.get("prefilter_signals", {}),
            "source_url": info.get("source_url"),
            "resolved_url": info.get("resolved_url"),
            "anchor_text": info.get("anchor_text"),
            "url_host": info.get("url_host"),
            "url_path": info.get("url_path"),
            "provider_family": info.get("provider_family"),
            "provider_expected_fields": info.get("provider_expected_fields"),
            "provider_group_key": info.get("provider_group_key"),
            "provider_candidate_urls": info.get("provider_candidate_urls"),
            "provider_recovered_fields": info.get("provider_recovered_fields"),
            "provider_recovery_status": info.get("provider_recovery_status"),
            "download_mode": info.get("download_mode"),
            "wrapper_detected": info.get("wrapper_detected"),
        }
        if extra:
            payload.update(extra)
        return {key: value for key, value in payload.items() if value is not None}

    def _is_controlled_truth_run(self):
        if not self._run_context.get("enabled"):
            return False

        run_id = str(self._run_context.get("run_id", "") or "").lower()
        if any(token in run_id for token in ("lockcheck", "regression", "frontend_full_run", "manual_frontend_run")):
            return True

        return bool(self._run_context.get("monitoring_dir"))

    def _should_gate_controlled_run_url(self, info):
        if not self._is_controlled_truth_run():
            return False
        if not info.get("is_url", False):
            return False

        provider_family = str(info.get("provider_family") or "").strip().lower()
        return not provider_family

    def _build_email_level_url_evidence_key(self, info):
        if not info or not info.get("is_url", False):
            return ""

        provider_family = str(info.get("provider_family") or "").strip().lower()
        if provider_family:
            return ""

        email_id = str(info.get("email_id") or "").strip()
        subject = str(info.get("subject") or "").strip().lower()
        sender = str(info.get("sender") or "").strip().lower()
        host = str(info.get("url_host") or "").strip().lower()

        if email_id:
            return f"email:{email_id}"
        if subject or sender:
            return f"fallback:{sender}|{subject}"
        if host:
            return f"host:{host}"
        return ""

    def _should_capture_email_level_url_evidence(self, info):
        aggregation_key = self._build_email_level_url_evidence_key(info)
        if not aggregation_key:
            return True, ""
        if aggregation_key in self._email_level_url_evidence_seen:
            return False, aggregation_key
        self._email_level_url_evidence_seen.add(aggregation_key)
        return True, aggregation_key

    def _effective_save_dir(self, requested_save_path):
        if self._run_context.get("enabled"):
            return self._run_context.get("output_dir", requested_save_path)
        return requested_save_path

    def _effective_date_range(self, date_from, date_to):
        if not self._run_context.get("enabled"):
            return date_from, date_to
        locked_from = self._run_context.get("locked_date_from", "") or date_from
        locked_to = self._run_context.get("locked_date_to", "") or date_to
        return locked_from, locked_to

    def _safe_write_run_config(self, email_address, auth_code="", api_key=""):
        if not self._run_context.get("enabled"):
            return
        self._diag_write_json(
            self._monitoring_path("run_config.json"),
            {
                **serialize_run_context(self._run_context),
                "run_id": self._current_run_id,
                "email_domain": self._packaged_diag_email_domain(email_address),
                "has_auth_code": bool(auth_code),
                "has_api_key": bool(api_key),
                "requested_save_path": self._requested_save_path,
                "effective_save_path": self._effective_save_path,
                "date_from": self._effective_date_from,
                "date_to": self._effective_date_to,
                "controlled_run": True,
                "stage_map": {
                    "start_processing": "active",
                    "frontend_processing_worker": "active",
                    "_run_processing_loop": "active",
                    "_simulate_processing": "legacy_inactive_not_invoked",
                    "cleanup_finalize": "active",
                },
            },
        )

    def _start_truth_audit_async(self, email_address, auth_code):
        if not self._run_context.get("enabled") or not email_address or not auth_code:
            return

        def _runner():
            try:
                module = importlib.import_module("audit_email_truth")
                collect_truth_table = getattr(module, "collect_truth_table")

                report = collect_truth_table(
                    email_address,
                    auth_code,
                    self._effective_date_from,
                    self._effective_date_to,
                )
                self._diag_write_json(self._monitoring_path("email_truth_audit.json"), report)
            except ModuleNotFoundError as exc:
                self._diag_write_json(
                    self._monitoring_path("email_truth_audit_error.json"),
                    {
                        "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                        "status": "skipped",
                        "reason": "AUDIT_EMAIL_TRUTH_MODULE_MISSING",
                        "detail": str(exc),
                    },
                )
            except Exception as exc:
                self._diag_write_json(
                    self._monitoring_path("email_truth_audit_error.json"),
                    {
                        "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
                        "error": str(exc),
                    },
                )

        threading.Thread(target=_runner, daemon=True).start()

    def _inspect_pdf_health(self, pdf_path):
        if not pdf_path or not str(pdf_path).lower().endswith(".pdf"):
            return None

        pdf_health = {
            "exists": False,
            "size_bytes": 0,
            "starts_with_pdf_magic": False,
            "fitz_open_ok": False,
            "page_count": 0,
            "first_page_text_len": 0,
            "render_to_base64_ok": None,
            "pdf_health_class": "ok",
            "pdf_health_reason": "",
        }
        try:
            pdf_health["exists"] = os.path.exists(pdf_path)
            if not pdf_health["exists"]:
                pdf_health["pdf_health_class"] = "unopenable_pdf"
                pdf_health["pdf_health_reason"] = "missing_file"
                return pdf_health

            pdf_health["size_bytes"] = os.path.getsize(pdf_path)
            if pdf_health["size_bytes"] <= 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = "zero_byte_file"
                return pdf_health

            try:
                with open(pdf_path, "rb") as fh:
                    pdf_health["starts_with_pdf_magic"] = fh.read(5).startswith(b"%PDF")
            except Exception:
                pass

            try:
                import fitz

                with fitz.open(pdf_path) as doc:
                    pdf_health["fitz_open_ok"] = True
                    pdf_health["page_count"] = len(doc)
                    if len(doc) > 0:
                        text_parts = []
                        for page_index in range(min(2, len(doc))):
                            text_parts.append(doc.load_page(page_index).get_text("text") or "")
                        pdf_health["first_page_text_len"] = len("".join(text_parts).strip())
            except Exception as exc:
                pdf_health["fitz_open_ok"] = False
                pdf_health["pdf_health_class"] = "corrupt_pdf" if not pdf_health["starts_with_pdf_magic"] else "unopenable_pdf"
                pdf_health["pdf_health_reason"] = str(exc)
                return pdf_health

            if pdf_health["page_count"] == 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = "no_pages"
            return pdf_health
        except Exception as exc:
            return {
                **pdf_health,
                "pdf_health_class": "unopenable_pdf",
                "pdf_health_reason": str(exc),
            }

    def _apply_render_health(self, pdf_health, base64_img):
        if not pdf_health:
            return None
        pdf_health = dict(pdf_health)
        pdf_health["render_to_base64_ok"] = bool(base64_img)
        if base64_img:
            return pdf_health
        if pdf_health["pdf_health_class"] == "ok":
            if pdf_health.get("page_count", 0) == 0 or pdf_health.get("size_bytes", 0) == 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = pdf_health.get("pdf_health_reason") or "no_renderable_pages"
            elif pdf_health.get("first_page_text_len", 0) == 0:
                pdf_health["pdf_health_class"] = "empty_pdf"
                pdf_health["pdf_health_reason"] = pdf_health.get("pdf_health_reason") or "empty_text_and_render_failed"
            else:
                pdf_health["pdf_health_class"] = "render_failed_pdf"
                pdf_health["pdf_health_reason"] = pdf_health.get("pdf_health_reason") or "base64_render_failed"
        return pdf_health

    @staticmethod
    def _compact_text(value):
        import re

        return re.sub(r"\s+", "", str(value or "")).strip().lower()

    def _extract_pdf_preview_text(self, pdf_path, max_pages=2):
        if not pdf_path or not str(pdf_path).lower().endswith(".pdf") or not os.path.exists(pdf_path):
            return ""
        try:
            import fitz

            texts = []
            with fitz.open(pdf_path) as doc:
                for page_index in range(min(max_pages, len(doc))):
                    texts.append(doc.load_page(page_index).get_text("text") or "")
            return "\n".join(texts).strip()
        except Exception:
            return ""

    def _has_structured_invoice_anchor(self, text):
        compact = self._compact_text(text)
        if not compact:
            return False
        anchors = [
            "发票号码",
            "发票代码",
            "购买方名称",
            "销售方名称",
            "价税合计",
            "开票日期",
            "invoice_number",
            "invoicenumber",
            "seller",
            "purchaser",
        ]
        return any(anchor in compact for anchor in anchors)

    def _match_wrapper_or_utility_reason(self, preview_text, info):
        compact = self._compact_text(preview_text)
        provider_family = str(info.get("provider_family", "") or "").lower()
        if not compact:
            return ""

        baiwang_hits = sum(
            1
            for token in ["发票预览", "下载pdf文件", "下载ofd文件", "下载xml文件", "关于百望", "previewinvoice"]
            if self._compact_text(token) in compact
        )
        if provider_family == "baiwang" and baiwang_hits >= 2 and not self._has_structured_invoice_anchor(preview_text):
            return "BAIWANG_WRAPPER_PAGE"

        if "connectingtotheitunesstore" in compact and "ifyoudonthaveitunes" in compact:
            return "UTILITY_PAGE_ITUNES_REDIRECT"

        if "票通云" in preview_text and (
            "提升运营效率" in preview_text or "发票链接一键" in preview_text or "多维链接" in preview_text
        ):
            return "UTILITY_PAGE_PROVIDER_MARKETING"

        return ""

    def _provider_fields_match(self, expected_fields, normalized_snapshot, info_json, recovered_fields=None):
        expected = dict(expected_fields or {})
        if not any(str(value or "").strip() for value in expected.values()):
            return True, "no_expected_fields"

        normalized_snapshot = normalized_snapshot or {}
        recovered_fields = recovered_fields or {}
        invoice_number = str(
            normalized_snapshot.get("InvoiceNumber")
            or info_json.get("InvoiceNumber")
            or info_json.get("invoice_number")
            or recovered_fields.get("invoice_number")
            or ""
        ).strip()
        seller = self._compact_text(
            normalized_snapshot.get("Seller")
            or info_json.get("Seller")
            or recovered_fields.get("seller")
            or ""
        )
        amount = str(
            normalized_snapshot.get("Amount")
            or info_json.get("Amount")
            or recovered_fields.get("amount")
            or ""
        ).strip()
        date = str(
            normalized_snapshot.get("Date")
            or info_json.get("Date")
            or recovered_fields.get("invoice_date")
            or ""
        ).strip()

        expected_number = str(expected.get("invoice_number") or "").strip()
        if expected_number:
            return invoice_number == expected_number, "invoice_number"

        expected_seller = self._compact_text(expected.get("seller") or "")
        expected_amount = str(expected.get("amount") or "").strip()
        expected_date = str(expected.get("invoice_date") or "").strip()

        seller_match = not expected_seller or expected_seller in seller or seller in expected_seller
        amount_match = not expected_amount or expected_amount == amount
        date_match = not expected_date or expected_date == date
        return seller_match and amount_match and date_match, "seller_amount_date"

    def _evaluate_document_acceptance(self, info, info_json, normalized_snapshot, pdf_health, pdf_path):
        preview_text = self._extract_pdf_preview_text(pdf_path)
        wrapper_reason = self._match_wrapper_or_utility_reason(preview_text, info)
        provider_family = str(info.get("provider_family", "") or "").lower()
        expected_fields = info.get("provider_expected_fields") or {}
        recovered_fields = info.get("provider_recovered_fields") or {}

        result = {
            "accepted": True,
            "reason_code": "",
            "bucket": "",
            "message": "",
            "provider_family": provider_family,
            "expected_fields": expected_fields,
            "pdf_preview_excerpt": preview_text[:500],
            "pdf_health_class": (pdf_health or {}).get("pdf_health_class", ""),
        }

        if wrapper_reason:
            result.update({
                "accepted": False,
                "reason_code": wrapper_reason,
                "bucket": "provider_wrapper_rejected",
                "message": "Downloaded result still looks like a provider wrapper or utility page.",
            })
            return result

        if provider_family == "baiwang":
            matched, matched_on = self._provider_fields_match(
                expected_fields,
                normalized_snapshot,
                info_json,
                recovered_fields=recovered_fields,
            )
            result["matched_on"] = matched_on
            if not matched:
                result.update({
                    "accepted": False,
                    "reason_code": "BAIWANG_EXPECTED_ENTITY_MISMATCH",
                    "bucket": "provider_entity_mismatch",
                    "message": "Downloaded Baiwang result does not match seller/amount/date/invoice anchors from email body.",
                })
                return result

        if provider_family in {"chinatax_direct_invoice", "bwjf_signed_invoice"}:
            invoice_number = str(
                normalized_snapshot.get("InvoiceNumber")
                or info_json.get("InvoiceNumber")
                or info_json.get("invoice_number")
                or recovered_fields.get("invoice_number")
                or ""
            ).strip()
            expected_number = str(expected_fields.get("invoice_number") or "").strip()
            if expected_number and invoice_number and invoice_number != expected_number:
                result.update({
                    "accepted": False,
                    "reason_code": "DIRECT_INVOICE_EXPECTED_ENTITY_MISMATCH",
                    "bucket": "provider_entity_mismatch",
                    "message": "Downloaded direct-invoice PDF exposes a conflicting invoice number.",
                    "matched_on": "invoice_number_conflict",
                })
                return result

            seller = self._compact_text(
                normalized_snapshot.get("Seller")
                or info_json.get("Seller")
                or recovered_fields.get("seller")
                or ""
            )
            expected_seller = self._compact_text(expected_fields.get("seller") or "")
            if expected_seller and seller and expected_seller not in seller and seller not in expected_seller:
                result.update({
                    "accepted": False,
                    "reason_code": "DIRECT_INVOICE_EXPECTED_ENTITY_MISMATCH",
                    "bucket": "provider_entity_mismatch",
                    "message": "Downloaded direct-invoice PDF exposes a conflicting seller entity.",
                    "matched_on": "seller_conflict",
                })
                return result

        return result

    def _set_run_state(self, run_state, status_text=None, progress=None, last_error=None):
        previous_state = self.run_state
        self.run_state = run_state
        self._is_running = run_state in {"running", "finalizing"}
        if status_text is not None:
            self.status_text = status_text
        if progress is not None:
            self.progress = progress
        if last_error is not None:
            self.last_error = last_error
        self._safe_emit_run_state_event(previous_state, run_state)

    def _begin_run(self, status_text):
        self.progress = 0
        self.logs = []
        self._stop_requested = False
        self.quota_exhausted = False
        self.quota_message = ""
        self._last_export_path = ""
        self.discovered_categories.clear()
        self.processed_invoices = []
        self.error_invoices = []
        self.stats = {"emails": 0, "invoices": 0, "errors": 0}
        self.audit_counts = {"manual_check": 0, "retention": 0, "raw_invoices": 0}
        self._email_level_url_evidence_seen = set()
        self._set_run_state("running", status_text=status_text, progress=0, last_error="")

    def _append_log(self, level, message, color="text-slate-700"):
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": level,
            "color": color,
            "msg": message,
        })

    def _request_safe_stop(self, message="正在安全停止，当前文件处理完后结束..."):
        if self._stop_requested:
            return
        self._stop_requested = True
        self.status_text = message
        self._append_log("停止", message, "text-amber-600")

    def _resolve_quota_message(self, error_text):
        normalized = str(error_text or "").lower()
        if not normalized:
            return ""
        quota_patterns = [
            "status code 402",
            "402 client error",
            "payment required",
            "余额不足",
            "额度不足",
            "quota",
            "insufficient balance",
            "insufficient_quota",
            "billing",
        ]
        if any(token in normalized for token in quota_patterns):
            return "GLM API 额度已耗尽，请充值或更换可用的 API Key。"
        return ""

    def _mark_quota_exhausted(self, message):
        self.quota_exhausted = True
        self.quota_message = message or "GLM API 额度已耗尽，请充值或更换可用的 API Key。"
        self.status_text = self.quota_message
        self._append_log("额度", self.quota_message, "text-rose-600")

    def _finish_run(self, success, status_text, last_error=""):
        if success:
            self._set_run_state("completed", status_text=status_text, progress=100, last_error="")
        else:
            failed_progress = self.progress if self.progress and self.progress < 100 else 99
            self._set_run_state("failed", status_text=status_text, progress=failed_progress, last_error=last_error)
        self._worker_thread = None

    def _legacy_start_async_finalizers_pre_release_prep(self, fetcher=None):
        def _runner():
            self._mark_finalizing()
            self._cleanup_temp_folders()
            if fetcher is not None:
                try:
                    fetcher.disconnect()
                except Exception as exc:
                    self.logs.append({
                        "time": time.strftime("[%H:%M:%S]"),
                        "type": "ERROR",
                        "color": "text-error",
                        "msg": f"Failed to disconnect mailbox cleanly: {exc}",
                    })

        cleanup_thread = threading.Thread(target=_runner, daemon=True)
        cleanup_thread.start()
        return cleanup_thread

    def _legacy_mark_finalizing_pre_release_prep(self):
        finalizing_progress = self.progress if self.progress >= 95 else 99
        self._set_run_state("finalizing", status_text="正在完成收尾...", progress=finalizing_progress)

    def _legacy_fail_run_pre_release_prep(self, status_text, error_message, fetcher=None, include_traceback=False):
        if fetcher is not None:
            self._start_async_finalizers(fetcher)

        if include_traceback:
            self._fail_run("澶勭悊寮傚父", error_msg, fetcher=fetcher, include_traceback=True)
            error_message = f"{error_message} | {traceback.format_exc()}"

        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "ERROR",
            "color": "text-error",
            "msg": f"系统发生异常: {error_message}",
        })
        self._finish_run(False, status_text, last_error=error_message)

    def _auto_start_local_scan(self):
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "INFO",
            "color": "text-slate-500",
            "msg": "Local auto scan is disabled in the release-prep candidate.",
        })
        return {"success": False, "message": "AUTO_LOCAL_SCAN_DISABLED"}

    def get_default_save_path(self):
        """前端初始化时获取默认的保存路径（桌面下的发票整理文件夹）"""
        run_context = self._refresh_run_context()
        if run_context.get("enabled"):
            return ensure_directory(run_context.get("output_dir", ""))
        return ensure_directory(resolve_default_save_path())

    def _normalize_user_save_path(self, path_value=""):
        run_context = self._refresh_run_context()
        if run_context.get("enabled"):
            return ensure_directory(run_context.get("output_dir", ""))

        candidate_path = str(path_value or "").strip()
        if not candidate_path:
            candidate_path = resolve_default_save_path()
        return ensure_directory(candidate_path)

    def _output_state_dir(self, save_path):
        return get_output_state_dir(save_path or self.get_default_save_path())

    def get_env_config(self):
        stored = self._settings_store.load() or {}
        return {
            "success": True,
            "email": str(stored.get("email", "") or ""),
            "auth_code": str(stored.get("auth_code", "") or ""),
            "api_key": str(stored.get("api_key", "") or ""),
        }

    def _default_user_settings(self):
        return {
            "email": "",
            "auth_code": "",
            "api_key": "",
            "save_path": self.get_default_save_path(),
            "date_from": "",
            "date_to": "",
            "quick_range": "last_30_days",
            "company": "",
            "remember_settings": True,
        }

    def load_user_settings(self):
        self._refresh_run_context()
        defaults = self._default_user_settings()
        stored = self._settings_store.load()

        merged = dict(defaults)
        merged.update({key: value for key, value in (stored or {}).items() if key in merged})
        merged["save_path"] = self._normalize_user_save_path(merged.get("save_path", ""))

        return {
            "success": True,
            "settings": merged,
            "settings_path": self._settings_store.settings_path,
        }

    def save_user_settings(self, settings):
        incoming = dict(settings or {})
        merged = self._default_user_settings()
        remember_settings = incoming.get("remember_settings", True)
        if remember_settings:
            merged.update({key: value for key, value in incoming.items() if key in merged})
            merged["remember_settings"] = True
        else:
            merged["remember_settings"] = False
        merged["save_path"] = self._normalize_user_save_path(merged.get("save_path", ""))
        self._settings_store.save(merged)
        return {"success": True, "message": "设置已保存", "path": self._settings_store.settings_path}

    def clear_user_settings(self):
        self._settings_store.clear()
        return {"success": True, "message": "本地设置已清除"}

    def get_run_context(self):
        return serialize_run_context(self._refresh_run_context())

    def test_connection(self, email, auth_code, api_key):

        """前端测试连接时调用（真实发包到大模型验证 Key）"""
        print(
            "Testing connection",
            {
                "email_domain": self._packaged_diag_email_domain(email),
                "has_auth_code": bool(auth_code),
                "api_key_length": len(api_key),
            },
        )
        
        try:
            import requests
        except ImportError:
            return {"success": False, "message": "连接失败 - 后端缺少 requests 依赖库"}
            
        # 邮箱连接测试（仅在填写授权码时执行）
        if auth_code:
            try:
                from email_fetcher import EmailFetcher
                channel = resolve_channel(email)
                fetcher = EmailFetcher(email, auth_code, imap_server=channel["imap_host"])
                if not fetcher.connect():
                    return {"success": False, "message": "邮箱 IMAP 登录验证失败"}
                fetcher.disconnect()
            except Exception as e:
                return {"success": False, "message": f"邮箱连接异常: {str(e)[:50]}"}
        
        if len(api_key) <= 5:
            return {"success": False, "message": "连接失败 - API Key 格式不正确"}
            
        try:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            # 构造一个最低成本的试探性请求（智谱 GLM API）
            payload = {
                "model": "glm-4-flash",
                "messages": [{"role": "user", "content": "Hi"}],
                "max_tokens": 5
            }
            response = requests.post(
                "https://open.bigmodel.cn/api/paas/v4/chat/completions",
                headers=headers,
                json=payload,
                timeout=15
            )
            
            if response.status_code == 200:
                return {"success": True, "message": "连接成功 - 智谱 GLM 服务已就绪"}
            elif response.status_code == 401:
                return {"success": False, "message": "连接失败 - API Key 鉴权未通过或无效"}
            elif response.status_code == 402:
                return {"success": False, "message": "连接失败 - GLM API 额度已耗尽，请充值或更换 API Key"}
            elif response.status_code == 429:
                return {"success": False, "message": "连接失败 - 触发限流，请求并发过高"}
            else:
                err_text = response.text
                quota_message = self._resolve_quota_message(f"{response.status_code} {err_text}")
                if quota_message:
                    return {"success": False, "message": f"连接失败 - {quota_message}"}
                print(f"GLM API Error HTTP {response.status_code}: {err_text}")
                return {"success": False, "message": f"连接异常 - 状态码: {response.status_code} - {err_text[:80]}"}
                
        except requests.exceptions.Timeout:
            return {"success": False, "message": "API连接失败 - 请求超时，请检查您的网络连接"}
        except requests.exceptions.ConnectionError:
            return {"success": False, "message": "API连接失败 - 无法连接到智谱 API 服务器"}
        except Exception as e:
            return {"success": False, "message": f"网络或API未知异常: {str(e)[:80]}"}

    def test_email_auth(self, email_address, auth_code):
        if not email_address or not auth_code:
            return {"success": False, "message": "请先填写邮箱地址和授权码"}
        import imaplib
        try:
            channel = resolve_channel(email_address)
            mail = imaplib.IMAP4_SSL(channel["imap_host"])
            mail.login(email_address, auth_code)
            mail.logout()
            return {"success": True, "message": "邮箱授权验证成功"}
        except Exception as e:
            return {"success": False, "message": f"验证失败，错误详情: {str(e)}"}

    def _simulate_processing(self, rules_text, save_path, date_from=None, date_to=None, email_address=None, auth_code=None, api_key=None):
        """后台线程: 连接邮箱抓取发票，执行 AI 处理与归档分类"""
        import os
        from email_fetcher import EmailFetcher
        from invoice_extractor import InvoiceExtractor
        
        fetcher = None
        self._begin_run("鍒濆鍖栧鐞嗗紩鎿?..")
        
        self.status_text = "初始化处理引擎..."
        range_msg = f"扫描时间: {date_from} 到 {date_to}" if date_from and date_to else "扫描时间: 默认近 30 天"
        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "信息:", "color": "text-blue-400", "msg": f"后端收到配置，正在初始化连接... ({range_msg})"})
        
        if not email_address or not auth_code or not api_key:
            self.status_text = "启动失败"
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": "未提供完整凭证(邮箱、授权码、或 API Key)，请返回设置。"})
            self._finish_run(False, "鍚姩澶辫触", last_error="MISSING_REQUIRED_CREDENTIALS")
            return
            
        try:
            # 1. 连接邮箱
            self.status_text = "正在连接邮箱服务..."
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "运行:", "color": "text-blue-400", "msg": f"正在登录邮箱 {email_address}..."})
            
            # 通过邮箱通道注册表选择 IMAP 服务器
            channel = resolve_channel(email_address)
            fetcher = EmailFetcher(email_address, auth_code, imap_server=channel["imap_host"], staging_dir="staging")
            if not fetcher.connect():
                self.status_text = "登录失败"
                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": "邮箱 IMAP 登录失败，请检查授权码或网络。"})
                self._finish_run(False, "鐧诲綍澶辫触", last_error="IMAP_LOGIN_FAILED")
                return
                
            # 2. 搜索邮件
            self.progress = 10
            self.status_text = "正在扫描邮件..."
            
            from datetime import datetime, timedelta
            # 处理时间范围
            since_date = date_from if date_from else (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            
            # IMAP 的 BEFORE 是非包含边界的。为了包含 date_to 当天，必须加一天 (exclusive boundary)
            if date_to:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
                before_date = dt_to.strftime("%Y-%m-%d")
            else:
                before_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
            
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "运行:", "color": "text-blue-400", "msg": f"开始检索 {since_date} 至 {before_date} 的邮件..."})
            
            email_ids = fetcher.fetch_emails_by_date(since_date=since_date, before_date=before_date)
            total_emails = len(email_ids)
            self.stats["emails"] = total_emails
            
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "成功:", "color": "text-emerald-400", "msg": f"共找到 {total_emails} 封邮件符合时间范围。"})
            
            if total_emails == 0:
                self._mark_finalizing()
                self.status_text = "扫描完成，无邮件"
                fetcher.disconnect()
                self._finish_run(True, "鎵弿瀹屾垚锛屾棤閭欢")
                self._is_running = False
                return
            
            # 3. 提取附件并保存到 staging
            self.progress = 30
            self.status_text = "正在下载发票附件..."
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "运行:", "color": "text-blue-400", "msg": "开始提取邮件附件及正文发票链接..."})
            
            attachments_info = fetcher.extract_attachments(email_ids)
            total_attachments = len(attachments_info)
            
            # 4. 初始化核心处理逻辑并执行
            self._run_processing_loop(attachments_info, api_key, save_path, since_date, before_date)
            
            # 在处理结束前执行清理
            self._cleanup_temp_folders()
            self._start_async_finalizers(fetcher)
            self._finish_run(True, "鎵瑰鐞嗗凡瀹屾垚")
            self.status_text = "批处理已完成"
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "完成:", "color": "text-emerald-400", "msg": f"扫描完成。总处理附件: {total_attachments}。"})
            self._is_running = False

        except Exception as e:
            error_msg = str(e) or "未知错误"
            import traceback
            full_traceback = traceback.format_exc()
            self.status_text = "处理异常"
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "ERROR", "color": "text-error", "msg": f"系统发生异常: {error_msg} | {full_traceback}"})
            traceback.print_exc()
            self.progress = 99
            self._is_running = False

    def _legacy_processing_worker_pre_release_prep(self, rules_text, save_path, date_from=None, date_to=None, email_address=None, auth_code=None, api_key=None):
        '''
        from email_fetcher import EmailFetcher

        fetcher = None
        self._begin_run("鍒濆鍖栧鐞嗗紩鎿?..")
        range_msg = f"鎵弿鏃堕棿: {date_from} 鍒?{date_to}" if date_from and date_to else "鎵弿鏃堕棿: 榛樿杩?30 澶?
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "淇℃伅:",
            "color": "text-blue-400",
            "msg": f"鍚庣鏀跺埌閰嶇疆锛屾鍦ㄥ垵濮嬪寲杩炴帴... ({range_msg})",
        })

        if not email_address or not auth_code or not api_key:
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "閿欒:",
                "color": "text-red-400",
                "msg": "鏈彁渚涘畬鏁村嚟璇?閭銆佹巿鏉冪爜銆佹垨 API Key)锛岃杩斿洖璁剧疆銆?",
            })
            self._finish_run(False, "鍚姩澶辫触", last_error="MISSING_REQUIRED_CREDENTIALS")
            return

        try:
            self._set_run_state("running", status_text="姝ｅ湪杩炴帴閭鏈嶅姟...")
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "杩愯:",
                "color": "text-blue-400",
                "msg": f"姝ｅ湪鐧诲綍閭 {email_address}...",
            })

            channel = resolve_channel(email_address)
            fetcher = EmailFetcher(email_address, auth_code, imap_server=channel["imap_host"], staging_dir="staging")
            if not fetcher.connect():
                self.logs.append({
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "閿欒:",
                    "color": "text-red-400",
                    "msg": "閭 IMAP 鐧诲綍澶辫触锛岃妫€鏌ユ巿鏉冪爜鎴栫綉缁溿€?",
                })
                self._finish_run(False, "鐧诲綍澶辫触", last_error="IMAP_LOGIN_FAILED")
                return

            self.progress = 10
            self.status_text = "姝ｅ湪鎵弿閭欢..."

            from datetime import datetime, timedelta
            since_date = date_from if date_from else (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            if date_to:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
                before_date = dt_to.strftime("%Y-%m-%d")
            else:
                before_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")

            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "杩愯:",
                "color": "text-blue-400",
                "msg": f"寮€濮嬫绱?{since_date} 鑷?{before_date} 鐨勯偖浠?..",
            })

            email_ids = fetcher.fetch_emails_by_date(since_date=since_date, before_date=before_date)
            total_emails = len(email_ids)
            self.stats["emails"] = total_emails
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "鎴愬姛:",
                "color": "text-emerald-400",
                "msg": f"鍏辨壘鍒?{total_emails} 灏侀偖浠剁鍚堟椂闂磋寖鍥淬€?",
            })

            if total_emails == 0:
                self._mark_finalizing()
                self._start_async_finalizers(fetcher)
                self._finish_run(True, "鎵弿瀹屾垚锛屾棤閭欢")
                return

            self.progress = 30
            self.status_text = "姝ｅ湪涓嬭浇鍙戠エ闄勪欢..."
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "杩愯:",
                "color": "text-blue-400",
                "msg": "寮€濮嬫彁鍙栭偖浠堕檮浠跺強姝ｆ枃鍙戠エ閾炬帴...",
            })

            attachments_info = fetcher.extract_attachments(email_ids)
            total_attachments = len(attachments_info)
            self._run_processing_loop(attachments_info, api_key, save_path, since_date, before_date)

            self._mark_finalizing()
            self._start_async_finalizers(fetcher)
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "瀹屾垚:",
                "color": "text-emerald-400",
                "msg": f"鎵弿瀹屾垚銆傛€诲鐞嗛檮浠? {total_attachments}銆?",
            })
            self._finish_run(True, "鎵瑰鐞嗗凡瀹屾垚")

        except Exception as e:
            error_msg = str(e) or "鏈煡閿欒"
            self._fail_run("澶勭悊寮傚父", error_msg, fetcher=fetcher, include_traceback=True)

        '''

    def _processing_worker(self, rules_text, save_path, date_from=None, date_to=None, email_address=None, auth_code=None, api_key=None):
        self._packaged_diag_write(
            "worker_enter",
            "_processing_worker",
            "success",
            summary={
                "effective_save_path": save_path,
                "date_from": date_from,
                "date_to": date_to,
                "email_domain": self._packaged_diag_email_domain(email_address),
                "has_auth_code": bool(auth_code),
                "has_api_key": bool(api_key),
            },
        )
        from datetime import datetime, timedelta
        from email_fetcher import EmailFetcher
        self._packaged_diag_write("worker_imports_ready", "_processing_worker", "success")

        fetcher = None
        self._begin_run("正在初始化任务...")
        self._packaged_diag_write("worker_begin_run_done", "_processing_worker", "success")
        self._safe_emit_stage_event(
            "frontend_processing_worker",
            "enter",
            {
                "requested_save_path": self._requested_save_path,
                "effective_save_path": save_path,
                "date_from": date_from,
                "date_to": date_to,
                **self._sensitive_summary(email_address, auth_code, api_key),
            },
        )
        range_msg = f"扫描范围: {date_from} -> {date_to}" if date_from and date_to else "扫描范围: 默认最近 30 天"
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "信息",
            "color": "text-blue-400",
            "msg": f"后端已接收请求，正在准备邮箱会话（{range_msg}）。",
        })

        if not email_address or not auth_code or not api_key:
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "错误",
                "color": "text-red-400",
                "msg": "缺少必要凭证：邮箱、授权码或 API Key。",
            })
            self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "failed", "reason": "missing_credentials"})
            self._finish_run(False, "启动失败", last_error="MISSING_REQUIRED_CREDENTIALS")
            return

        try:
            self._packaged_diag_write(
                "before_progress_gt_5",
                "_processing_worker",
                "success",
                summary={"target_progress": 10},
            )
            self._set_run_state("running", status_text="正在连接邮箱...", progress=10)
            self._packaged_diag_write("after_progress_10", "_processing_worker", "success")
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "运行",
                "color": "text-blue-400",
                "msg": f"正在连接邮箱 {email_address}...",
            })

            channel = resolve_channel(email_address)
            fetcher = EmailFetcher(
                email_address,
                auth_code,
                imap_server=channel["imap_host"],
                staging_dir=self._run_context.get("staging_dir") or "staging",
                monitoring_dir=self._run_context.get("monitoring_dir"),
            )
            self._packaged_diag_write("network_connect_before", "_processing_worker", "success")
            connect_result = fetcher.connect()
            self._packaged_diag_write(
                "network_connect_after",
                "_processing_worker",
                "success" if connect_result else "failure",
                summary={"connect_result": bool(connect_result)},
            )
            if not connect_result:
                self.logs.append({
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "错误",
                    "color": "text-red-400",
                    "msg": "邮箱登录失败，请检查授权码和 IMAP 设置。",
                })
                self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "failed", "reason": "imap_login_failed"})
                self._finish_run(False, "邮箱登录失败", last_error="IMAP_LOGIN_FAILED")
                return

            self._set_run_state("running", status_text="正在扫描邮件...", progress=20)
            since_date = date_from if date_from else (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            if date_to:
                before_date = (datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            else:
                before_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")

            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "运行",
                "color": "text-blue-400",
                "msg": f"正在扫描时间范围 {since_date} -> {before_date} 的邮件。",
            })

            self._packaged_diag_write("network_scan_before", "_processing_worker", "success")
            email_ids = fetcher.fetch_emails_by_date(since_date=since_date, before_date=before_date)
            self._packaged_diag_write(
                "network_scan_after",
                "_processing_worker",
                "success",
                summary={"scan_result_count": len(email_ids)},
            )
            total_emails = len(email_ids)
            self.stats["emails"] = total_emails
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "信息",
                "color": "text-emerald-400",
                "msg": f"共匹配到 {total_emails} 封邮件。",
            })

            if self._stop_requested:
                self._mark_finalizing()
                self._start_async_finalizers(fetcher)
                self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "stopped_before_extract", "emails": total_emails})
                self._finish_run(True, "已安全停止")
                return

            if total_emails == 0:
                self._mark_finalizing()
                self._start_async_finalizers(fetcher)
                self.logs.append({
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "完成",
                    "color": "text-emerald-400",
                    "msg": "未找到符合条件的邮件，任务已结束。",
                })
                self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "completed", "emails": total_emails, "attachments": 0})
                self._finish_run(True, "处理完成")
                return

            self._set_run_state("running", status_text="正在提取附件...", progress=30)
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "运行",
                "color": "text-blue-400",
                "msg": "正在下载并提取附件...",
            })

            self._packaged_diag_write("extract_attachments_before", "_processing_worker", "success")
            attachments_info = fetcher.extract_attachments(email_ids)
            self._packaged_diag_write(
                "extract_attachments_after",
                "_processing_worker",
                "success",
                summary={"extract_result_count": len(attachments_info)},
            )
            total_attachments = len(attachments_info)
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "信息",
                "color": "text-emerald-400",
                "msg": f"共提取到 {total_attachments} 个附件。",
            })

            if self._stop_requested:
                self._mark_finalizing()
                self._start_async_finalizers(fetcher)
                self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "stopped_after_extract", "emails": total_emails, "attachments": total_attachments})
                self._finish_run(True, "已安全停止")
                return

            self._run_processing_loop(attachments_info, api_key, save_path, since_date, before_date, rules_text)

            if self.quota_exhausted:
                self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "failed", "reason": self.quota_message or "QUOTA_EXHAUSTED"})
                self._fail_run("GLM API 额度不足", self.quota_message or "GLM API 额度已耗尽，请充值或更换可用的 API Key。", fetcher=fetcher)
                return

            if self._stop_requested:
                self._mark_finalizing()
                self._start_async_finalizers(fetcher)
                self.logs.append({
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "完成",
                    "color": "text-amber-600",
                    "msg": "已完成安全停止，已处理结果已保留。",
                })
                self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "stopped", "emails": total_emails, "attachments": total_attachments})
                self._finish_run(True, "已安全停止")
                return

            # CWT 取消撮合后处理: 找到匹配的预订确认单，移入 Manual_Check
            self._cwt_cancellation_matching(save_path)

            self._mark_finalizing()
            self._start_async_finalizers(fetcher)
            self.logs.append({
                "time": time.strftime("[%H:%M:%S]"),
                "type": "完成",
                "color": "text-emerald-400",
                "msg": f"处理循环已完成，共处理 {total_attachments} 个附件。",
            })
            self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "completed", "emails": total_emails, "attachments": total_attachments})
            self._finish_run(True, "处理完成")
        except Exception as exc:
            self._packaged_diag_write("worker_exception", "_processing_worker", "exception", exc=exc)
            self._safe_emit_stage_event("frontend_processing_worker", "exit", {"result": "failed", "reason": str(exc) or "UNKNOWN_ERROR"})
            self._fail_run("处理失败", str(exc) or "UNKNOWN_ERROR", fetcher=fetcher, include_traceback=True)

    def _run_processing_loop(self, attachments_info, api_key, save_path, since_date=None, before_date=None, rules_text=""):
        self._packaged_diag_write(
            "run_loop_enter",
            "_run_processing_loop",
            "success",
            summary={"attachments": len(attachments_info or [])},
        )
        import copy
        import os
        import traceback
        from invoice_extractor import InvoiceExtractor
        
        total_attachments = len(attachments_info)
        if total_attachments == 0:
            self._safe_emit_stage_event("_run_processing_loop", "enter", {"attachments": 0, "save_path": save_path})
            self._safe_emit_stage_event("_run_processing_loop", "exit", {"result": "completed", "attachments": 0})
            return

        self._safe_emit_stage_event(
            "_run_processing_loop",
            "enter",
            {
                "attachments": total_attachments,
                "save_path": save_path,
                "since_date": since_date,
                "before_date": before_date,
            },
        )
            
        extractor = InvoiceExtractor(api_key=api_key, output_dir=save_path)
        output_state_dir = self._output_state_dir(save_path)
        extractor.processed_records_file = os.path.join(output_state_dir, "processed_records.json")
        self._packaged_diag_write("run_loop_extractor_ready", "_run_processing_loop", "success")
        
        # --- PHASE 1: RECONCILIATION GROUND TRUTH ---
        ground_truth_files = {info['filepath']: info for info in attachments_info}
        processed_filepaths = set()
        
        # Load business logic deduplication records
        business_records = extractor.load_processed_records()
        trace_store = self._create_document_trace_store(
            output_path=self._run_context.get("debug_trace_path") or None
        )
        phase2_completed = False
        phase2_had_error = False
        loop_result = "completed"
        browser_first_recorded = False

        def _build_normalized_fields(fields):
            if not fields or not isinstance(fields, dict):
                return None

            raw_date = str(fields.get("Date", ""))
            clean_date = raw_date.replace("/", "").replace("-", "").replace("年", "").replace("月", "").replace("日", "").strip()
            raw_amount = str(fields.get("Amount", ""))
            clean_amount = raw_amount.replace(",", "").replace("¥", "").replace("￥", "").replace("元", "").replace(" ", "").strip()

            return {
                "Date": clean_date if clean_date else "未知",
                "Amount": clean_amount if clean_amount else "未知",
                "Purchaser": str(fields.get("Purchaser", "")),
                "Seller": str(fields.get("Seller", "")),
                "Type": str(fields.get("Type", "")),
                "InvoiceCode": str(fields.get("InvoiceCode", "")).strip(),
                "InvoiceNumber": str(fields.get("InvoiceNumber", "")).strip(),
                "is_invoice": fields.get("is_invoice", True),
            }

        def _mark_combine_not_applicable(document_id, reason_code, message):
            trace_store.set_fields(
                document_id,
                combine_keys={"status": "not_applicable", "reason_code": reason_code},
                combine_result={
                    "status": "not_applicable",
                    "reason_code": reason_code,
                    "message": message,
                },
            )

        def _record_combine_candidate(document_id, combine_type, role, meta):
            if not document_id:
                return

            trace_store.set_fields(
                document_id,
                combine_keys={
                    "combine_type": combine_type,
                    "document_role": role,
                    "date": meta.get("date", ""),
                    "amount": meta.get("amount", ""),
                    "seller": meta.get("seller", ""),
                    "filename": meta.get("filename", ""),
                    "reason_code": f"{combine_type.upper()}_COMBINE_CANDIDATE",
                },
            )

        def _record_combine_result(document_id, status, reason_code, message=None, **extras):
            if not document_id:
                return

            payload = {"status": status, "reason_code": reason_code}
            if message:
                payload["message"] = message
            payload.update(extras)
            trace_store.set_fields(document_id, combine_result=payload)

        def _build_history_key(info, file_name, pdf_path):
            import hashlib

            legacy_key = hashlib.md5(
                f"{info.get('subject', '')}_{file_name}_{info.get('tier', 0)}".encode("utf-8")
            ).hexdigest()
            if info.get("is_url", False):
                return f"url:{legacy_key}"

            try:
                with open(pdf_path, "rb") as source_file:
                    file_digest = hashlib.sha256(source_file.read()).hexdigest()
                return f"att:{file_digest}"
            except Exception:
                return f"att-legacy:{legacy_key}"

        def _finalize_trace_defaults():
            for record in trace_store.iter_records():
                document_id = record["document_id"]
                archive_target = record.get("archive_target")
                archive_folder = os.path.basename(os.path.dirname(archive_target)) if archive_target else ""
                combine_applicable = archive_folder in {"打车", "住宿发票"}

                if record.get("combine_keys") is None:
                    if combine_applicable:
                        trace_store.set_fields(
                            document_id,
                            combine_keys={
                                "status": "pending",
                                "reason_code": "COMBINE_KEYS_NOT_RECORDED",
                            },
                        )
                    else:
                        trace_store.set_fields(
                            document_id,
                            combine_keys={
                                "status": "not_applicable",
                                "reason_code": "COMBINE_NOT_APPLICABLE",
                            },
                        )

                if record.get("combine_result") is None:
                    if combine_applicable:
                        trace_store.set_fields(
                            document_id,
                            combine_result={
                                "status": "not_evaluated" if (phase2_had_error or not phase2_completed) else "not_matched",
                                "reason_code": "COMBINE_NOT_EVALUATED" if (phase2_had_error or not phase2_completed) else "COMBINE_NO_MATCH",
                                "message": "Combine stage did not emit a document-specific result.",
                            },
                        )
                    else:
                        trace_store.set_fields(
                            document_id,
                            combine_result={
                                "status": "not_applicable",
                                "reason_code": "COMBINE_NOT_APPLICABLE",
                                "message": "Document did not enter a combine-enabled archive folder.",
                            },
                        )

        try:
            # 处理每个附件
            success_count = 0
            processed_provider_groups = set()
            for i, info in enumerate(attachments_info):
                if self._stop_requested:
                    loop_result = "stopped"
                    self._append_log("停止", "已收到停止指令，当前文件处理完毕后结束本次运行。", "text-amber-600")
                    break
                if not self._is_running:
                    break
                    
                source_filename = os.path.basename(info['filepath'])
                file_name = source_filename
                pdf_path = info['filepath']
                tier_info = info.get('tier', '未知')
                
                try:
                    time.sleep(0.2) # 确保磁盘 IO 已完成
                    print(f">>> [{i+1}/{total_attachments}] 开始处理文件: {file_name}")
                    import hashlib
                    document_id = hashlib.md5(
                        f"{info.get('filepath', '')}|{info.get('subject', '')}|{file_name}|{info.get('tier', 0)}|{i}".encode("utf-8")
                    ).hexdigest()
                    trace_store.start_document(
                        source_filename=source_filename,
                        source_path=info.get("filepath"),
                        document_id=document_id,
                    )
                    prefilter_reason_code = info.get("prefilter_reason_code")
                    prefilter_metadata = self._attachment_diag_metadata(
                        info,
                        file_name=file_name,
                        document_id=document_id,
                        extra={
                            "tier": info.get("tier", 0),
                            "candidate_index": i + 1,
                        },
                    )
                    if info.get("candidate_action") == "retain_only":
                        should_capture_evidence = True
                        aggregation_key = ""
                        if info.get("is_url", False):
                            should_capture_evidence, aggregation_key = self._should_capture_email_level_url_evidence(info)
                        if not should_capture_evidence:
                            trace_store.set_fields(
                                document_id,
                                naming_result={"status": "skipped", "reason_code": prefilter_reason_code or "EMAIL_LEVEL_URL_EVIDENCE_AGGREGATED"},
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Email-level URL evidence was already retained for this email, so duplicate low-confidence URLs were collapsed.",
                            )
                            trace_store.record_failure_event(
                                document_id,
                                prefilter_reason_code or "EMAIL_LEVEL_URL_EVIDENCE_AGGREGATED",
                                "prefilter",
                                "Duplicate low-confidence URL was collapsed into existing email-level audit evidence.",
                                severity="skipped",
                            )
                            processed_filepaths.add(pdf_path)
                            continue
                        if aggregation_key:
                            prefilter_metadata = dict(prefilter_metadata)
                            prefilter_metadata.update(
                                {
                                    "email_level_url_evidence": True,
                                    "email_level_url_evidence_key": aggregation_key,
                                }
                            )
                        retained_path = self._retain_artifact(
                            save_path,
                            pdf_path,
                            "prefilter_b_retained",
                            "P0 预过滤判定为 B 层 durable retention 候选",
                            prefilter_metadata,
                        )
                        self.stats["errors"] += 1
                        self.logs.append({
                            "time": time.strftime("[%H:%M:%S]"),
                            "type": "保全:",
                            "color": "text-blue-400",
                            "msg": f"前置过滤 B 层候选已保全: {os.path.basename(retained_path)}",
                        })
                        self.error_invoices.append({
                            "id": f"inv_prefilter_{time.time()}_{i}",
                            "date": "---",
                            "amount": "---",
                            "category": "预过滤保全",
                            "merchant": "高/中置信度候选",
                            "path": retained_path,
                            "name": os.path.basename(retained_path),
                            "sColor": "bg-blue-500",
                            "status": "已保全待判断",
                            "reason": prefilter_reason_code or "P0_B_RETENTION",
                            "rColor": "text-blue-600 border-blue-200 bg-blue-50",
                        })
                        trace_store.set_fields(
                            document_id,
                            naming_result={"status": "skipped", "reason_code": prefilter_reason_code or "P0_B_RETENTION"},
                            archive_target=retained_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Prefilter B-layer retain-only candidate did not enter archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            prefilter_reason_code or "P0_B_RETENTION",
                            "prefilter",
                            "前置过滤判定为 B 层 durable retention 候选。",
                            severity="fallback",
                        )
                        processed_filepaths.add(pdf_path)
                        continue

                    if info.get("candidate_action") == "manual_review":
                        review_path = self._send_to_manual_check(
                            save_path,
                            pdf_path,
                            prefilter_reason_code or "P0_C_MANUAL_REVIEW",
                            metadata=prefilter_metadata,
                            is_url=info.get("is_url", False),
                        )
                        self.stats["errors"] += 1
                        self.logs.append({
                            "time": time.strftime("[%H:%M:%S]"),
                            "type": "复核:",
                            "color": "text-yellow-400",
                            "msg": f"前置过滤 C 层候选已进入 Manual_Check: {os.path.basename(review_path)}",
                        })
                        self.error_invoices.append({
                            "id": f"inv_prefilter_{time.time()}_{i}",
                            "date": "---",
                            "amount": "---",
                            "category": "人工复核",
                            "merchant": "低置信度候选",
                            "path": review_path,
                            "name": os.path.basename(review_path),
                            "sColor": "bg-yellow-500",
                            "status": "待人工复核",
                            "reason": prefilter_reason_code or "P0_C_MANUAL_REVIEW",
                            "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50",
                        })
                        trace_store.set_fields(
                            document_id,
                            naming_result={"status": "skipped", "reason_code": prefilter_reason_code or "P0_C_MANUAL_REVIEW"},
                            archive_target=review_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Prefilter C-layer candidate was routed to Manual_Check before archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            prefilter_reason_code or "P0_C_MANUAL_REVIEW",
                            "prefilter",
                            "前置过滤判定为 C 层人工复核候选。",
                            severity="fallback",
                        )
                        processed_filepaths.add(pdf_path)
                        continue

                    # --- State Persistence Check ---
                    # 生成唯一 ID 以防重复提交大模型
                    history_key = _build_history_key(info, file_name, pdf_path)
                    
                    history_file_path = os.path.join(output_state_dir, ".antigravity_history.json")
                    processed_history = set()
                    if os.path.exists(history_file_path):
                        try:
                            import json
                            with open(history_file_path, "r", encoding="utf-8") as hf:
                                processed_history = set(json.load(hf))
                        except Exception:
                            pass
                            
                    if history_key in processed_history:
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "记忆:", "color": "text-blue-400", "msg": f"跳过已成功归档文件: {file_name}"})
                        self.progress = 50 + int(((i + 1) / total_attachments) * 45)
                        retained_path = None
                        if not info.get('is_url', False):
                            retained_path = self._retain_artifact(
                                save_path,
                                pdf_path,
                                "history_skipped",
                                "命中历史去重记录，未重复提交模型",
                                {
                                    "subject": info.get("subject", ""),
                                    "tier": info.get("tier", 0),
                                    "history_key": history_key
                                }
                            )
                            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "保全:", "color": "text-blue-400", "msg": f"已保留历史跳过副本: {os.path.basename(retained_path)}"})
                        trace_store.set_fields(
                            document_id,
                            naming_result={"status": "skipped", "reason_code": "HISTORY_DUPLICATE_SKIP"},
                            archive_target=retained_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Document was skipped by history de-duplication before archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            "HISTORY_DUPLICATE_SKIP",
                            "history",
                            "命中历史去重记录，跳过本次处理。",
                            severity="skipped",
                        )
                        processed_filepaths.add(pdf_path)
                        continue
                        
                    self.progress = 50 + int(((i + 1) / total_attachments) * 45)  # 50% -> 95%
                    self.status_text = f"正在解析发票 (第 {i+1} 个，共 {total_attachments} 个)..."
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "解析:", "color": "text-purple-400", "msg": f"AI 正在识别: {file_name}"})
                    
                    # --- 网页链接发票无头下载支持 (Link Download) ---
                    if info.get('is_url', False):
                        provider_group_key = str(info.get("provider_group_key", "") or "")
                        if provider_group_key and provider_group_key in processed_provider_groups:
                            trace_store.set_fields(
                                document_id,
                                naming_result={"status": "skipped", "reason_code": "PROVIDER_GROUP_ALREADY_PROCESSED"},
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Provider recovery group was already handled by an earlier candidate.",
                            )
                            processed_filepaths.add(pdf_path)
                            continue

                        if self._should_gate_controlled_run_url(info):
                            should_capture_evidence, aggregation_key = self._should_capture_email_level_url_evidence(info)
                            if not should_capture_evidence:
                                trace_store.set_fields(
                                    document_id,
                                    naming_result={"status": "skipped", "reason_code": "CONTROLLED_RUN_NON_PROVIDER_URL_AGGREGATED"},
                                )
                                _mark_combine_not_applicable(
                                    document_id,
                                    "COMBINE_NOT_APPLICABLE",
                                    "Controlled truth run collapsed duplicate non-provider URLs into a single email-level audit record.",
                                )
                                trace_store.record_failure_event(
                                    document_id,
                                    "CONTROLLED_RUN_NON_PROVIDER_URL_AGGREGATED",
                                    "execution_gate",
                                    "Duplicate non-provider URL was collapsed into an existing email-level controlled-run audit record.",
                                    severity="skipped",
                                )
                                processed_filepaths.add(pdf_path)
                                continue
                            gate_reason_code = "CONTROLLED_RUN_NON_PROVIDER_URL_SKIPPED"
                            gate_message = (
                                "Controlled truth run skipped non-provider URL execution to preserve audit evidence "
                                "and avoid blocking the main chain on non-target webpages."
                            )
                            gate_metadata = dict(prefilter_metadata)
                            gate_metadata.update(
                                {
                                    "prefilter_reason_code": gate_reason_code,
                                    "original_prefilter_reason_code": prefilter_reason_code or "",
                                    "controlled_run_url_gate": True,
                                    "controlled_run_url_gate_reason": gate_message,
                                    "email_level_url_evidence": True,
                                    "email_level_url_evidence_key": aggregation_key,
                                }
                            )
                            retained_path = self._retain_artifact(
                                save_path,
                                info.get("source_url") or pdf_path,
                                "controlled_run_non_provider_url",
                                gate_message,
                                gate_metadata,
                            )
                            self.stats["errors"] += 1
                            self.logs.append(
                                {
                                    "time": time.strftime("[%H:%M:%S]"),
                                    "type": "淇濆叏:",
                                    "color": "text-blue-400",
                                    "msg": f"鍙楁帶璺戞壒宸蹭繚鍏ㄩ潪 provider URL锛?{os.path.basename(retained_path)}",
                                }
                            )
                            self.error_invoices.append(
                                {
                                    "id": f"inv_controlled_url_{time.time()}_{i}",
                                    "date": "---",
                                    "amount": "---",
                                    "category": "鍙楁帶璺戞壒淇濆叏",
                                    "merchant": "闈?provider URL",
                                    "path": retained_path,
                                    "name": os.path.basename(retained_path),
                                    "sColor": "bg-blue-500",
                                    "status": "宸蹭繚鍏ㄥ緟鍒ゆ柇",
                                    "reason": gate_reason_code,
                                    "rColor": "text-blue-600 border-blue-200 bg-blue-50",
                                }
                            )
                            trace_store.set_fields(
                                document_id,
                                naming_result={"status": "skipped", "reason_code": gate_reason_code},
                                archive_target=retained_path,
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Controlled truth run URL gate skipped non-provider URL execution before archive/combine.",
                            )
                            trace_store.record_failure_event(
                                document_id,
                                gate_reason_code,
                                "execution_gate",
                                gate_message,
                                severity="fallback",
                            )
                            processed_filepaths.add(pdf_path)
                            continue

                        if not browser_first_recorded:
                            self._packaged_diag_write(
                                "browser_first_before",
                                "_run_processing_loop",
                                "success",
                                summary={"file_name": file_name},
                            )
                        from pdf_converter import PDFConverter
                        converter = PDFConverter(
                            staging_dir=self._run_context.get("staging_dir") or "staging",
                            timeout_ms=30000,
                        )
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "抓取:", "color": "text-blue-400", "msg": f"正在启动无头浏览器抓取网页: {file_name[:30]}"})
                        try:
                            link_results = converter.process_invoice_links(
                                pdf_path,
                                info.get('subject', 'Link_Invoice'),
                                f"url_{i}",
                                return_metadata=True,
                                candidate_info=info,
                            )
                            if not browser_first_recorded:
                                self._packaged_diag_write(
                                    "browser_first_after",
                                    "_run_processing_loop",
                                    "success" if bool(link_results) else "failure",
                                    summary={
                                        "file_name": file_name,
                                        "url_result_count": len(link_results or []),
                                    },
                                )
                                browser_first_recorded = True
                        except Exception as exc:
                            if not browser_first_recorded:
                                self._packaged_diag_write(
                                    "browser_first_exception",
                                    "_run_processing_loop",
                                    "exception",
                                    summary={"file_name": file_name},
                                    exc=exc,
                                )
                                browser_first_recorded = True
                            raise
                        if not link_results:
                            self.stats["errors"] += 1
                            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"抓取发票链接失败或遇反爬验证码: {file_name}"})
                            trace_store.set_fields(
                                document_id,
                                naming_result={"status": "skipped", "reason_code": "URL_DOWNLOAD_FAILED"},
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Document failed before archive stage during URL download.",
                            )
                            trace_store.record_failure_event(
                                document_id,
                                "URL_DOWNLOAD_FAILED",
                                "source_download",
                                "抓取发票链接失败或遇反爬验证码。",
                                severity="failure",
                            )
                            processed_filepaths.add(pdf_path) # [Fix] Mark URL links as processed on failure
                            continue
                        link_result = dict(link_results[0] or {})
                        if provider_group_key:
                            processed_provider_groups.add(provider_group_key)

                        link_status = str(link_result.get("status") or "").strip().lower()
                        if link_status in {"failed", "skipped"} and not link_result.get("pdf_path"):
                            link_reason_code = str(link_result.get("reason_code") or "URL_DOWNLOAD_FAILED").strip() or "URL_DOWNLOAD_FAILED"
                            link_message = (
                                str(link_result.get("message") or "").strip()
                                or str(link_result.get("failure_stage") or "").strip()
                                or "URL processing terminated before a usable PDF entered the main chain."
                            )
                            is_runtime_filtered = link_reason_code == "URL_NON_INVOICE_PAGE_SKIPPED"
                            bucket_suffix = "url_runtime_filtered" if is_runtime_filtered else "url_runtime_failed"
                            retained_path = self._retain_artifact(
                                save_path,
                                link_result.get("resolved_url") or info.get("source_url") or pdf_path,
                                bucket_suffix,
                                link_message,
                                self._attachment_diag_metadata(
                                    info,
                                    file_name=file_name,
                                    document_id=document_id,
                                    extra={
                                        "tier": tier_info,
                                        "url_runtime_result": link_result,
                                    },
                                ),
                            )
                            self.stats["errors"] += 1
                            if is_runtime_filtered:
                                self.logs.append(
                                    {
                                        "time": time.strftime("[%H:%M:%S]"),
                                        "type": "保全:",
                                        "color": "text-blue-400",
                                        "msg": f"运行时过滤了非票据页面，已保留证据: {os.path.basename(retained_path)}",
                                    }
                                )
                                self.error_invoices.append(
                                    {
                                        "id": f"inv_url_filtered_{time.time()}_{i}",
                                        "date": "---",
                                        "amount": "---",
                                        "category": "保留记录",
                                        "merchant": "运行时过滤",
                                        "path": retained_path,
                                        "name": os.path.basename(retained_path),
                                        "sColor": "bg-blue-500",
                                        "status": "已保全待判断",
                                        "reason": link_reason_code,
                                        "rColor": "text-blue-600 border-blue-200 bg-blue-50",
                                    }
                                )
                                failure_severity = "fallback"
                            else:
                                self.logs.append(
                                    {
                                        "time": time.strftime("[%H:%M:%S]"),
                                        "type": "错误:",
                                        "color": "text-red-400",
                                        "msg": f"链接抓取失败，已保留证据: {os.path.basename(retained_path)}",
                                    }
                                )
                                self.error_invoices.append(
                                    {
                                        "id": f"inv_url_failed_{time.time()}_{i}",
                                        "date": "---",
                                        "amount": "---",
                                        "category": "解析失败",
                                        "merchant": "链接抓取",
                                        "path": retained_path,
                                        "name": os.path.basename(retained_path),
                                        "sColor": "bg-yellow-500",
                                        "status": "处理异常",
                                        "reason": link_reason_code,
                                        "rColor": "text-red-600 border-red-200 bg-red-50",
                                    }
                                )
                                failure_severity = "failure"
                            trace_store.set_fields(
                                document_id,
                                source_download_result=link_result,
                                naming_result={"status": "skipped", "reason_code": link_reason_code},
                                archive_target=retained_path,
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "URL runtime handling terminated before archive/combine.",
                            )
                            trace_store.record_failure_event(
                                document_id,
                                link_reason_code,
                                "source_download",
                                link_message,
                                severity=failure_severity,
                            )
                            processed_filepaths.add(pdf_path)
                            continue

                        if link_result.get("status") == "provider_recovery_failed":
                            provider_family = str(link_result.get("provider_family") or info.get("provider_family") or "").lower()
                            bucket_suffix = str(link_result.get("retention_bucket_suffix") or provider_family or "unknown_provider").strip().strip("\\/")
                            if not bucket_suffix:
                                bucket_suffix = "unknown_provider"
                            recovery_reason_code = link_result.get("reason_code") or f"PROVIDER_RECOVERY_FAILED_{provider_family.upper() or 'UNKNOWN'}"
                            recovery_message = (
                                link_result.get("provider_recovery_message")
                                or f"{provider_family or 'Provider'} recovery failed before a confirmed invoice PDF could enter the main chain."
                            )
                            retained_path = self._retain_artifact(
                                save_path,
                                info.get("source_url") or pdf_path,
                                os.path.join("provider_recovery_failed", bucket_suffix),
                                recovery_message,
                                self._attachment_diag_metadata(
                                    info,
                                    file_name=file_name,
                                    document_id=document_id,
                                    extra={
                                        "tier": tier_info,
                                        "provider_recovery": link_result,
                                    },
                                ),
                            )
                            self.stats["errors"] += 1
                            self.logs.append({
                                "time": time.strftime("[%H:%M:%S]"),
                                "type": "淇濆叏:",
                                "color": "text-red-400",
                                "msg": f"百望恢复失败，已保全审计对象: {os.path.basename(retained_path)}",
                            })
                            trace_store.set_fields(
                                document_id,
                                source_download_result=link_result,
                                naming_result={"status": "skipped", "reason_code": recovery_reason_code},
                                archive_target=retained_path,
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Provider recovery failed before archive stage.",
                            )
                            trace_store.record_failure_event(
                                document_id,
                                recovery_reason_code,
                                "source_download",
                                recovery_message,
                                severity="failure",
                            )
                            processed_filepaths.add(pdf_path)
                            continue

                        info = dict(info)
                        info.update({
                            "resolved_url": link_result.get("resolved_url", ""),
                            "download_mode": link_result.get("download_mode", ""),
                            "wrapper_detected": link_result.get("wrapper_detected", False),
                            "provider_family": link_result.get("provider_family", info.get("provider_family", "")),
                            "provider_recovered_fields": link_result.get("selected_fields", {}),
                            "provider_recovery_status": link_result.get("status", ""),
                        })
                        trace_store.set_fields(document_id, source_download_result=link_result)
                        pdf_path = link_result["pdf_path"]
                        file_name = os.path.basename(pdf_path)
                        time.sleep(0.5) # 确保无头浏览器下载的 PDF 磁盘 IO 写入完成

                    pdf_health = None
                    if str(pdf_path).lower().endswith(".pdf"):
                        pdf_health = self._inspect_pdf_health(pdf_path)
                        if pdf_health:
                            trace_store.set_fields(document_id, pdf_health=pdf_health)
                    
                    # Step A: 转图片
                    base64_img = extractor.pdf_to_base64_image(pdf_path)
                    pdf_health = self._apply_render_health(pdf_health, base64_img)
                    if pdf_health:
                        trace_store.set_fields(document_id, pdf_health=pdf_health)
                    if not base64_img:
                        self.stats["errors"] += 1
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"无法读取附件图像: {file_name}"})
                        self._record_error_log(save_path, info.get('subject', file_name), "PDF损坏或不是标准发票格式")
                        retained_path = self._retain_artifact(
                            save_path,
                            pdf_path,
                            "unreadable",
                            "本地预处理失败，无法转成可供识别的图像",
                            self._attachment_diag_metadata(
                                info,
                                file_name=file_name,
                                document_id=document_id,
                                extra={
                                    "tier": tier_info,
                                    "pdf_health": pdf_health,
                                },
                            ),
                        )
                        self.error_invoices.append({
                            "id": f"inv_{time.time()}_{i}",
                            "date": "---",
                            "amount": "---",
                            "category": "预处理失败",
                            "merchant": "无法读取附件",
                            "path": retained_path,
                            "name": os.path.basename(retained_path),
                            "sColor": "bg-yellow-500",
                            "status": "待人工复核",
                            "reason": "无法读取附件图像，已保全原件",
                            "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50"
                        })
                        trace_store.set_fields(
                            document_id,
                            naming_result={"status": "skipped", "reason_code": "PDF_TO_IMAGE_FAILED"},
                            archive_target=retained_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Document failed during image conversion before archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            "PDF_TO_IMAGE_FAILED",
                            "preprocess",
                            "无法读取附件图像。",
                            severity="failure",
                        )
                        processed_filepaths.add(pdf_path) # [Fix] Mark as processed on failure
                        continue
                        
                    # Step B: 前置成本控制 (Cost Filter)
                    # 利用本地 PyMuPDF 先查一次文本里的年份
                    cost_gate_valid = True
                    try:
                        import fitz
                        temp_doc = fitz.open(pdf_path)
                        full_text = ""
                        for p in range(min(2, len(temp_doc))):
                            full_text += temp_doc.load_page(p).get_text()
                        temp_doc.close()
                        
                        start_year = since_date.split('-')[0] if since_date else ""
                        end_year = before_date.split('-')[0] if before_date else ""
                        
                        # 如果文本中有绝对不是当前年份的强烈特征，且没有当前年份，可以考虑阻断 (此处采取保守策略，如果文本为空则放行图文API)
                        if full_text.strip():
                            if start_year and end_year and start_year == end_year:
                                if "20" in full_text and start_year not in full_text:
                                    # Very loose check: if there's a 20xx but our year isn't there, might be an old invoice. 
                                    pass # For safety, we rely on the final Date Gatekeeper, but this is the hook for future strict OCR filtering
                    except Exception:
                        pass
                        
                    # Step C: LLM 信息提取
                    try:
                        info_json = extractor.extract_info_via_llm(base64_img, custom_rules=rules_text, pdf_path=pdf_path)
                    except Exception as extraction_error:
                        quota_message = self._resolve_quota_message(extraction_error)
                        if quota_message:
                            raise QuotaExceededError(quota_message) from extraction_error
                        raise
                    extraction_trace = copy.deepcopy(getattr(extractor, "last_extraction_trace", {}) or {})
                    trace_store.set_fields(document_id, extractor_raw_result=extraction_trace or None)
                    if extraction_trace.get("reason_code") == "TRACK_A_FAILED_TRACK_B_FALLBACK":
                        trace_store.record_failure_event(
                            document_id,
                            "TRACK_A_FAILED_TRACK_B_FALLBACK",
                            "extraction",
                            extraction_trace.get("track_a", {}).get("message"),
                            severity="fallback",
                        )
                    if not info_json:
                        self.stats["errors"] += 1
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"大模型全部引擎提取失败，移入人工区: {file_name}"})
                        self._record_error_log(save_path, info.get('subject', file_name), "大模型未能识别出标准的 JSON 数据")
                        
                        # 直接抛给 route_and_rename_file 进行 Manual_Check 兜底保存
                        manual_success, result_path = extractor.route_and_rename_file(pdf_path, None)
                        naming_trace = copy.deepcopy(getattr(extractor, "last_route_trace", {}) or {})
                        if not manual_success:
                            result_path = self._retain_artifact(
                                save_path,
                                pdf_path,
                                "model_failed_fallback",
                                "模型彻底失败且人工区复制失败，保全原件",
                                {
                                    "subject": info.get("subject", ""),
                                    "tier": tier_info,
                                    "file_name": file_name
                                }
                            )
                        elif result_path:
                            self.audit_counts["manual_check"] += 1
                            self._safe_emit_artifact_event(
                                "manual_check",
                                result_path,
                                document_id=document_id,
                                source_kind=info.get("source_kind"),
                                reason_code=naming_trace.get("reason_code") or "ROUTE_TO_MANUAL_CHECK",
                                category="Manual_Check",
                                extra=self._attachment_diag_metadata(
                                    info,
                                    file_name=file_name,
                                    document_id=document_id,
                                ),
                            )
                        trace_store.set_fields(
                            document_id,
                            classification_result={"status": "failed", "reason_code": "EXTRACTOR_ALL_ENGINES_FAILED"},
                            naming_result=naming_trace or {"status": "failed", "reason_code": "ROUTE_TO_MANUAL_CHECK"},
                            archive_target=result_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Document did not reach combine stage after extraction failure.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            "EXTRACTOR_ALL_ENGINES_FAILED",
                            "extraction",
                            "大模型未能识别出标准的 JSON 数据。",
                            severity="failure",
                        )
                        if naming_trace.get("reason_code"):
                            trace_store.record_failure_event(
                                document_id,
                                naming_trace["reason_code"],
                                "naming",
                                naming_trace.get("error_message") or "解析失败后进入 Manual_Check 兜底。",
                                severity="fallback" if naming_trace.get("used_manual_check") else "failure",
                            )
                        processed_filepaths.add(pdf_path)
                        self.error_invoices.append({
                            "id": f"inv_{time.time()}_{i}", "date": "---", "amount": "---", "category": "人工复核",
                            "merchant": "模型提取彻底失败", "path": result_path, "name": os.path.basename(result_path),
                            "sColor": "bg-yellow-500", "status": "待人工复核", "reason": "解析失败移入 Manual_Check",
                            "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50"
                        })
                        continue
                        
                    # 行程单/水单豁免: 先识别文件名关键词，判断为行程单/水单则强制豁免 is_invoice 过滤
                    doc_type_peek = str(info_json.get("Type", ""))
                    normalized_snapshot = _build_normalized_fields(info_json)
                    acceptance_check = self._evaluate_document_acceptance(
                        info,
                        info_json,
                        normalized_snapshot,
                        pdf_health,
                        pdf_path,
                    )
                    trace_store.set_fields(document_id, document_acceptance=acceptance_check)
                    if not acceptance_check.get("accepted", True):
                        retained_path = self._retain_artifact(
                            save_path,
                            pdf_path,
                            acceptance_check.get("bucket") or "provider_guard_rejected",
                            acceptance_check.get("message") or "Downloaded result rejected by document acceptance gate.",
                            self._attachment_diag_metadata(
                                info,
                                file_name=file_name,
                                document_id=document_id,
                                extra={
                                    "tier": tier_info,
                                    "pdf_health": pdf_health,
                                    "normalized_snapshot": normalized_snapshot,
                                    "document_acceptance": acceptance_check,
                                },
                            ),
                        )
                        self.stats["errors"] += 1
                        self.logs.append({
                            "time": time.strftime("[%H:%M:%S]"),
                            "type": "保全:",
                            "color": "text-yellow-400",
                            "msg": f"下载结果被文档验收闸门拦截: {os.path.basename(retained_path)}",
                        })
                        trace_store.set_fields(
                            document_id,
                            normalized_fields=normalized_snapshot,
                            classification_result={
                                "status": "rejected",
                                "reason_code": acceptance_check.get("reason_code"),
                                "provider_family": acceptance_check.get("provider_family", ""),
                            },
                            naming_result={"status": "skipped", "reason_code": acceptance_check.get("reason_code")},
                            archive_target=retained_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Document was rejected by the provider/entity acceptance gate before archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            acceptance_check.get("reason_code") or "DOCUMENT_ACCEPTANCE_REJECTED",
                            "document_acceptance",
                            acceptance_check.get("message") or "Downloaded result rejected by document acceptance gate.",
                            severity="failure",
                        )
                        processed_filepaths.add(pdf_path)
                        continue
                    is_itinerary_or_folio = (
                        any(kw in file_name for kw in ["行程单", "行程报销单", "报销单", "folio", "Folio"])
                        or any(kw in doc_type_peek for kw in ["行程单", "报销单", "水单", "folio", "Folio"])
                    )

                    is_invoice = info_json.get("is_invoice", True)
                    if not is_itinerary_or_folio and (str(is_invoice).lower() == "false" or is_invoice is False):
                        retained_path = self._retain_artifact(
                            save_path,
                            pdf_path,
                            "model_rejected",
                            "模型判定为非票据，转人工复核保全",
                            {
                                "subject": info.get("subject", ""),
                                "tier": tier_info,
                                "file_name": file_name,
                                "rejection_reason": info_json.get("rejection_reason", ""),
                                "model_type": info_json.get("Type", "")
                            }
                        )
                        self.stats["errors"] += 1
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "保全:", "color": "text-yellow-400", "msg": f"模型拒绝文件已转人工复核保全: {os.path.basename(retained_path)}"})
                        self.error_invoices.append({
                            "id": f"inv_{time.time()}_{i}",
                            "date": info_json.get("Date", "---"),
                            "amount": f"¥ {info_json.get('Amount', '0.00')}",
                            "category": "模型拒绝",
                            "merchant": info_json.get("Seller", "未知开票方"),
                            "path": retained_path,
                            "name": os.path.basename(retained_path),
                            "sColor": "bg-yellow-500",
                            "status": "待人工复核",
                            "reason": info_json.get("rejection_reason", "模型判定为非票据"),
                            "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50"
                        })
                        trace_store.set_fields(
                            document_id,
                            normalized_fields=normalized_snapshot,
                            classification_result={
                                "status": "rejected",
                                "reason_code": "MODEL_REJECTED_NOT_INVOICE",
                                "model_type": info_json.get("Type", ""),
                                "is_itinerary_or_folio": is_itinerary_or_folio,
                                "rejection_reason": info_json.get("rejection_reason", ""),
                            },
                            naming_result={"status": "skipped", "reason_code": "MODEL_REJECTED_NOT_INVOICE"},
                            archive_target=retained_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Document was rejected as non-invoice before archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            "MODEL_REJECTED_NOT_INVOICE",
                            "classification",
                            info_json.get("rejection_reason", "模型判定为非票据。"),
                            severity="failure",
                        )
                        processed_filepaths.add(pdf_path) # Mark as intentionally discarded
                        continue

                    clean_date = normalized_snapshot.get("Date", "") if normalized_snapshot else ""

                    # 日期合法性校验：发票年份不得超过邮件搜索结束日期的年份 (防止截图等被误归档)
                    if clean_date and len(clean_date) >= 4:
                        try:
                            year = int(clean_date[:4])
                            # 允许范围: 2000 ~ 搜索结束年份
                            max_year = int(before_date[:4]) if before_date and len(before_date) >= 4 else 2025
                            print(f">>> [年份校验] file={file_name}, date={clean_date}, year={year}, before_date={before_date}, max_year={max_year}")
                            if year > max_year or year < 2000:
                                self.logs.append({
                                    "time": time.strftime("[%H:%M:%S]"),
                                    "type": "标记:",
                                    "color": "text-yellow-400",
                                    "msg": f"日期异常 ({clean_date})，已记录 DATE_OUT_OF_RANGE 标记并继续主链路",
                                })
                                trace_store.set_fields(
                                    document_id,
                                    normalized_fields=normalized_snapshot,
                                    classification_result={
                                        "status": "flagged",
                                        "reason_code": "DATE_OUT_OF_RANGE",
                                        "detected_date": clean_date,
                                        "max_year": max_year,
                                    },
                                )
                                trace_store.record_failure_event(
                                    document_id,
                                    "DATE_OUT_OF_RANGE",
                                    "normalization",
                                    f"年份超出合法范围 (≤{max_year})。",
                                    severity="fallback",
                                )
                        except (ValueError, TypeError):
                            pass

                    info_json["Date"] = normalized_snapshot.get("Date", "未知") if normalized_snapshot else "未知"
                    info_json["Amount"] = normalized_snapshot.get("Amount", "未知") if normalized_snapshot else "未知"

                    purchaser = str(info_json.get("Purchaser", ""))
                    doc_type = str(info_json.get("Type", ""))
                    seller = str(info_json.get("Seller", ""))

                    # 1. 通用分类归一化 (LLM Type + 文件名 + 商户名 三重检查)
                    original_doc_type = doc_type
                    classification_reason_codes = []
                    
                    # 1a. 国旅运通 (CWT) 优先检测 — 必须在通用分类之前
                    #     多信号检测: 发件人(@citsgbt.com) / Seller / Subject / 文件名
                    #     支持转发场景: notification@citsgbt.com 被转发时 sender 非原始
                    #     GBT Travel Services 是国旅运通的海外出票实体（SCCT号发票→差旅服务费）
                    #     取消知会: 文件名含"取消" → 标记 _cwt_cancellation，后续撮合送 Manual_Check
                    _is_cwt = (
                        "citsgbt.com" in str(info.get("sender", "")).lower()
                        or any(kw in seller for kw in ["国旅运通", "CWT", "Carlson Wagonlit", "citsgbt", "GBT Travel"])
                        or any(kw in str(info.get("subject", "")).lower() for kw in ["citsgbt", "国旅运通", "cwt", "cits gbt", "scct"])
                        or any(kw in file_name.lower() for kw in ["citsgbt", "国旅运通", "cwt", "scct"])
                    )
                    if _is_cwt:
                        if "取消" in file_name:
                            doc_type = "住宿确认单"
                            classification_reason_codes.append("CWT_HOTEL_CANCELLATION")
                            info_json["_cwt_cancellation"] = True
                        # SCCT/GBT Travel 发票 → 差旅服务费（最高优先级，不受 LLM type 干扰）
                        elif any(kw in seller for kw in ["GBT Travel"]) or "scct" in file_name.lower():
                            doc_type = "差旅服务费"
                            classification_reason_codes.append("CLASSIFIED_AS_CWT_SERVICE_FEE")
                        elif any(kw in file_name.lower() for kw in ["flight", "air", "机票", "航班", "行程单 - 机票"]):
                            doc_type = "航班行程单"
                            classification_reason_codes.append("CLASSIFIED_AS_CWT_FLIGHT_BY_FILENAME")
                        elif any(kw in doc_type.lower() for kw in ["机票", "航班", "flight", "air"]):
                            doc_type = "航班行程单"
                            classification_reason_codes.append("CLASSIFIED_AS_CWT_FLIGHT")
                        elif any(kw in file_name.lower() for kw in ["酒店", "行程单 - 酒店"]):
                            doc_type = "住宿确认单"
                            classification_reason_codes.append("CLASSIFIED_AS_CWT_HOTEL_BY_FILENAME")
                        else:
                            doc_type = "住宿确认单"
                            classification_reason_codes.append("CLASSIFIED_AS_CWT_HOTEL")
                    # 1b. 行程单识别 — 区分航班行程单 vs 打车行程单
                    elif any(kw in doc_type for kw in ["行程单", "报销单"]) or any(kw in file_name for kw in ["行程单", "行程报销单", "报销单"]):
                        # "行程单 - 机票" 或 doc_type/seller 含航空关键词 → 航班行程单
                        _fn_lower = file_name.lower()
                        _is_flight = (
                            "机票" in _fn_lower
                            or any(kw in doc_type.lower() for kw in ["机票", "航班", "flight", "air"])
                            or any(kw in seller for kw in ["航空", "Airlines", "Air China", "东航", "南航", "国航"])
                        )
                        if _is_flight:
                            doc_type = "航班行程单"
                            classification_reason_codes.append("CLASSIFIED_AS_FLIGHT_ITINERARY")
                        else:
                            doc_type = "打车"
                            info_json["_is_itinerary"] = True
                            classification_reason_codes.append("CLASSIFIED_AS_RIDE_ITINERARY")
                    # 1c. 打车发票 (LLM Type 或 Seller 中含网约车关键词)
                    elif any(kw in doc_type for kw in ["打车", "出租", "滴滴", "高德", "约车"]):
                        doc_type = "打车"
                        classification_reason_codes.append("CLASSIFIED_AS_RIDE_BY_TYPE")
                    elif any(kw in seller for kw in ["滴滴", "高德", "约车", "盛智", "畅行"]):
                        doc_type = "打车"  # Seller 兜底: LLM Type 不准时用商户名补救
                        classification_reason_codes.append("CLASSIFIED_AS_RIDE_BY_SELLER")
                    # 1d. 火车票
                    elif any(kw in doc_type for kw in ["火车", "高铁", "铁路"]):
                        doc_type = "火车票"
                        classification_reason_codes.append("CLASSIFIED_AS_TRAIN_BY_TYPE")
                    # 1e. 水单
                    elif any(kw in doc_type for kw in ["水单", "Folio", "账单", "folio"]):
                        doc_type = "住宿发票"
                        info_json["_is_folio"] = True
                        classification_reason_codes.append("CLASSIFIED_AS_HOTEL_FOLIO")
                    elif "住宿" in doc_type:
                        doc_type = "住宿发票"
                        classification_reason_codes.append("CLASSIFIED_AS_HOTEL_INVOICE")
                    else:
                        classification_reason_codes.append("CLASSIFICATION_FROM_MODEL_TYPE")
                    info_json["Type"] = doc_type

                    # 2. 豁免白名单 (使用 document_types 注册表)
                    from document_types import is_exempt_type
                    is_exempt = is_exempt_type(doc_type)

                    # 3. 公司抬头严格拦截网 (仅针对非豁免票据，如餐饮、住宿发票等)
                    if not is_exempt:
                        from company_rules import is_company_purchaser
                        active_company = self._run_context.get("company") or (self._settings_store.load() or {}).get("company") or ""
                        if not is_company_purchaser(purchaser, active_company):
                            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "拦截:", "color": "text-yellow-400", "msg": f"非公司抬头 ({purchaser}), 已移入隔离区"})
                            info_json["Type"] = "个人非报销发票"
                            classification_reason_codes.append("CLASSIFIED_AS_PERSONAL_NON_REIMBURSEMENT")

                    # Step C: 回传记录前端
                    category_name = info_json.get("category", info_json.get("Type", "未知分类"))
                    self.discovered_categories.add(category_name)
                    trace_store.set_fields(
                        document_id,
                        normalized_fields=normalized_snapshot,
                        classification_result={
                            "status": "classified",
                            "original_type": original_doc_type,
                            "final_type": info_json.get("Type", ""),
                            "category": category_name,
                            "is_invoice": is_invoice,
                            "is_exempt": is_exempt,
                            "is_itinerary_or_folio": is_itinerary_or_folio,
                            "reason_code": classification_reason_codes[-1] if classification_reason_codes else "CLASSIFICATION_FROM_MODEL_TYPE",
                            "reason_codes": classification_reason_codes,
                        },
                    )
                    
                    # --- Business Logic Deduplication (拦截重复发票) ---
                    invoice_code = info_json.get("InvoiceCode", "").strip()
                    invoice_number = info_json.get("InvoiceNumber", "").strip()
                    if invoice_code or invoice_number:
                        if extractor.is_duplicate(invoice_code, invoice_number, business_records):
                            self.stats["errors"] += 1
                            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "去重:", "color": "text-yellow-400", "msg": f"已触发发票去重机制 (发票代码/号码相同): {file_name}"})
                            processed_filepaths.add(pdf_path)
                            duplicate_path = self._retain_artifact(
                                save_path,
                                pdf_path,
                                "duplicates",
                                "命中业务去重规则，保留原件供复核",
                                {
                                    "subject": info.get("subject", ""),
                                    "tier": tier_info,
                                    "file_name": file_name,
                                    "invoice_code": invoice_code,
                                    "invoice_number": invoice_number
                                }
                            )
                            
                            # Log as duplicate in error output for visibility
                            self.error_invoices.append({
                                "id": f"inv_{time.time()}_{i}",
                                "date": info_json.get("Date", "---"),
                                "amount": f"¥ {info_json.get('Amount', '0.00')}",
                                "category": category_name,
                                "merchant": info_json.get("Seller", "未知开票方"),
                                "path": duplicate_path,
                                "name": file_name,
                                "sColor": "bg-gray-400",
                                "status": "重复跳过",
                                "reason": f"代码:{invoice_code} 号码:{invoice_number}",
                                "rColor": "text-gray-600 border-gray-200 bg-gray-50"
                            })
                            trace_store.set_fields(
                                document_id,
                                naming_result={"status": "skipped", "reason_code": "BUSINESS_DUPLICATE_SKIPPED"},
                                archive_target=duplicate_path,
                            )
                            _mark_combine_not_applicable(
                                document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Document was skipped by business de-duplication before archive/combine.",
                            )
                            trace_store.record_failure_event(
                                document_id,
                                "BUSINESS_DUPLICATE_SKIPPED",
                                "dedup",
                                f"发票代码/号码重复: {invoice_code}/{invoice_number}",
                                severity="skipped",
                            )
                            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "保全:", "color": "text-gray-400", "msg": f"重复票据已保留原件副本: {file_name}"})
                            continue
                    
                    # Step D: 分发与降级逻辑 (Tier 3 阻断)
                    # CWT 豁免: 已通过多信号检测的 CWT 文档不受 Tier 3 降级
                    # 这些文档有明确的发件人/Subject/文件名信号，分类可信度高
                    _cwt_classified = any(
                        rc.startswith("CLASSIFIED_AS_CWT_") or rc.startswith("CWT_")
                        for rc in classification_reason_codes
                    )
                    if tier_info == 3 and not _cwt_classified:
                        self.stats["errors"] += 1
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "异常:", "color": "text-yellow-400", "msg": f"边缘触发需人工复核: {file_name}"})
                        retained_path = self._send_to_manual_check(
                            save_path,
                            pdf_path,
                            "TIER3_MANUAL_REVIEW",
                            metadata={
                                "subject": info.get("subject", ""),
                                "tier": tier_info,
                                "file_name": file_name,
                            },
                            is_url=False,
                        )
                        self.error_invoices.append({
                            "id": f"inv_{time.time()}_{i}",
                            "date": info_json.get("Date", "---"),
                            "amount": f"¥ {info_json.get('Amount', '0.00')}",
                            "category": category_name,
                            "merchant": info_json.get("Seller", "未知开票方"),
                            "path": retained_path,
                            "name": os.path.basename(retained_path),
                            "sColor": "bg-yellow-500",
                            "status": "待人工复核",
                            "reason": "边缘触发",
                            "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50"
                            # 重点：Tier 3 不执行 route_and_rename_file 归档，保留在暂存区或直接标记
                        })
                        trace_store.set_fields(
                            document_id,
                            naming_result={"status": "skipped", "reason_code": "TIER3_MANUAL_REVIEW"},
                            archive_target=retained_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Tier 3 documents do not enter archive/combine.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            "TIER3_MANUAL_REVIEW",
                            "classification",
                            "Tier 3 边缘触发，转人工复核。",
                            severity="failure",
                        )
                        processed_filepaths.add(pdf_path)
                        continue
                    
                    # 正常成功提取 (Tier 1, 2, 4)
                    
                    # --- Date Gatekeeper 已移除 ---
                    # 日期过滤仅由 IMAP 邮件搜索阶段控制 (SINCE/BEFORE)
                    # 发票开票日期不应被限制，因为开票日期可能早于邮件日期

                    # CWT 取消知会 → 直接送 Manual_Check，记录撮合信息
                    if info_json.get("_cwt_cancellation"):
                        retained_path = self._send_to_manual_check(
                            save_path, pdf_path, "CWT_HOTEL_CANCELLATION",
                            metadata={"subject": info.get("subject", ""), "file_name": file_name},
                        )
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-amber-400", "msg": f"CWT取消知会已送人工复核: {file_name}"})
                        if not hasattr(self, '_cwt_cancellation_registry'):
                            self._cwt_cancellation_registry = []
                        self._cwt_cancellation_registry.append({
                            "file_name": file_name,
                            "manual_check_path": retained_path,
                        })
                        trace_store.set_fields(document_id, archive_target=retained_path)
                        processed_filepaths.add(pdf_path)
                        continue

                    # 构建归档路由规则: CWT 类型映射到正确的目标文件夹
                    # 航班行程单→机票, 住宿确认单→住宿发票(与发票/水单排序挨着)
                    from document_types import get_archive_folder
                    _archive_rules = {}
                    _cur_type = info_json.get("Type", "")
                    _mapped_folder = get_archive_folder(_cur_type)
                    if _mapped_folder != _cur_type:
                        _archive_rules[_cur_type] = _mapped_folder
                    success, result_path = extractor.route_and_rename_file(pdf_path, info_json, custom_rules=_archive_rules or None)
                    naming_trace = copy.deepcopy(getattr(extractor, "last_route_trace", {}) or {})
                    
                    if success:
                        success_count += 1
                        self.stats["invoices"] = success_count
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "成功:", "color": "text-emerald-400", "msg": f"[{category_name}] 归档至: {os.path.basename(result_path)}"})
                        trace_store.set_fields(
                            document_id,
                            naming_result=naming_trace or {"status": "archived", "reason_code": None},
                            archive_target=result_path,
                        )
                        self._safe_emit_artifact_event(
                            "archive",
                            result_path,
                            document_id=document_id,
                            source_kind=info.get("source_kind"),
                            reason_code=naming_trace.get("reason_code"),
                            category=category_name,
                            extra=self._attachment_diag_metadata(
                                info,
                                file_name=file_name,
                                document_id=document_id,
                                extra={
                                    "final_type": info_json.get("Type", ""),
                                    "seller": info_json.get("Seller", ""),
                                },
                            ),
                        )
                        
                        # 同步到前端数据分析模块要用到的发票列
                        self.processed_invoices.append({
                            "id": f"inv_{time.time()}_{i}",
                            "date": info_json.get("Date", "---"),
                            "amount": f"¥ {info_json.get('Amount', '0.00')}",
                            "category": category_name,
                            "merchant": info_json.get("Seller", "未知开票方"),
                            "path": result_path
                        })
                        processed_filepaths.add(pdf_path)
                        
                        # 记录成功提取的指纹，避免重复扣费
                        processed_history.add(history_key)
                        try:
                            import json
                            with open(history_file_path, "w", encoding="utf-8") as hf:
                                json.dump(list(processed_history), hf)
                        except Exception:
                            pass
                            
                        # 更新 Business 去重字典并持久化
                        if invoice_code or invoice_number:
                            dup_key = f"{invoice_code}_{invoice_number}"
                            business_records[dup_key] = {"file": os.path.basename(result_path), "date": info_json.get("Date", ""), "amount": info_json.get("Amount", "")}
                            extractor.save_processed_records(business_records)
                    else:
                        self.stats["errors"] += 1
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "跳过:", "color": "text-yellow-400", "msg": f"未归档或放入人工分类: {file_name} ({result_path})"})
                        self._record_error_log(save_path, info.get('subject', file_name), result_path)
                        original_error = result_path
                        if "Manual_Check" not in result_path:
                            result_path = self._retain_artifact(
                                save_path,
                                pdf_path,
                                "archive_failures",
                                "归档阶段失败，保全原件供后续复核",
                                {
                                    "subject": info.get("subject", ""),
                                    "tier": tier_info,
                                    "file_name": file_name,
                                    "archive_error": result_path
                                }
                            )
                        else:
                            self.audit_counts["manual_check"] += 1
                            self._safe_emit_artifact_event(
                                "manual_check",
                                result_path,
                                document_id=document_id,
                                source_kind=info.get("source_kind"),
                                reason_code=naming_trace.get("reason_code") or "ROUTE_TO_MANUAL_CHECK",
                                category="Manual_Check",
                                extra=self._attachment_diag_metadata(
                                    info,
                                    file_name=file_name,
                                    document_id=document_id,
                                ),
                            )
                        trace_store.set_fields(
                            document_id,
                            naming_result=naming_trace or {"status": "failed", "reason_code": "ARCHIVE_COPY_FAILED"},
                            archive_target=result_path,
                        )
                        _mark_combine_not_applicable(
                            document_id,
                            "COMBINE_NOT_APPLICABLE",
                            "Archive stage failed before combine could run.",
                        )
                        trace_store.record_failure_event(
                            document_id,
                            naming_trace.get("reason_code", "ARCHIVE_STAGE_FAILED"),
                            "naming",
                            naming_trace.get("error_message") or original_error,
                            severity="failure",
                        )
                        processed_filepaths.add(pdf_path) # [Fix] Always mark as processed
                        
                        # 纳入人工分类的即使失败也要视为已处理，由 Manual_Check 承接
                        if "Manual_Check" in result_path:
                            processed_filepaths.add(pdf_path)
                            self.error_invoices.append({
                                "id": f"inv_{time.time()}_{i}",
                                "date": "---",
                                "amount": "---",
                                "category": "人工复核",
                                "merchant": "无法自动分类",
                                "path": result_path,
                                "name": os.path.basename(result_path),
                                "sColor": "bg-yellow-500",
                                "status": "待人工复核",
                                "reason": result_path,
                                "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50"
                            })
                        else:
                            self.error_invoices.append({
                                "id": f"inv_{time.time()}_{i}",
                                "date": info_json.get("Date", "---"),
                                "amount": f"¥ {info_json.get('Amount', '0.00')}",
                                "category": category_name,
                                "merchant": info_json.get("Seller", "未知开票方"),
                                "path": result_path,
                                "name": os.path.basename(result_path),
                                "sColor": "bg-yellow-500",
                                "status": "归档失败待复核",
                                "reason": "归档阶段失败，已保全原件",
                                "rColor": "text-yellow-600 border-yellow-200 bg-yellow-50"
                            })

                except Exception as loop_err:
                    if isinstance(loop_err, QuotaExceededError):
                        self.stats["errors"] += 1
                        err_msg = str(loop_err)
                        self._mark_quota_exhausted(err_msg)
                        retained_path = self._retain_artifact(
                            save_path,
                            pdf_path,
                            "glm_quota_exhausted",
                            err_msg,
                            {
                                "subject": info.get("subject", ""),
                                "tier": info.get("tier", 0),
                                "file_name": file_name,
                                "error": err_msg,
                            },
                        )
                        self.error_invoices.append({
                            "id": f"inv_{time.time()}_{i}",
                            "date": "---",
                            "amount": "---",
                            "category": "额度不足/服务异常",
                            "merchant": "GLM API",
                            "path": retained_path,
                            "name": os.path.basename(retained_path),
                            "sColor": "bg-red-500",
                            "status": "额度不足/服务异常",
                            "reason": err_msg,
                            "rColor": "text-red-600 border-red-200 bg-red-50"
                        })
                        trace_store.set_fields(document_id, archive_target=retained_path)
                        trace_store.record_failure_event(
                            document_id,
                            "GLM_QUOTA_EXHAUSTED",
                            "extraction",
                            err_msg,
                            severity="failure",
                        )
                        processed_filepaths.add(pdf_path)
                        loop_result = "quota_exhausted"
                        break

                    self.stats["errors"] += 1
                    err_msg = str(loop_err)
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"处理单张发票失败: {file_name} - {err_msg}"})
                    retained_path = self._retain_artifact(
                        save_path,
                        pdf_path,
                        "processing_errors",
                        "处理单张票据时抛出异常，转人工复核保全",
                        {
                            "subject": info.get("subject", ""),
                            "tier": info.get("tier", 0),
                            "file_name": file_name,
                            "error": err_msg
                        }
                    )
                    self.error_invoices.append({
                        "id": f"inv_{time.time()}_{i}",
                        "date": "---",
                        "amount": "---",
                        "category": "解析失败",
                        "merchant": "系统错误",
                        "path": retained_path,
                        "name": os.path.basename(retained_path),
                        "sColor": "bg-yellow-500",
                        "status": "处理异常",
                        "reason": err_msg,
                        "rColor": "text-red-600 border-red-200 bg-red-50"
                    })
                    trace_store.set_fields(document_id, archive_target=retained_path)
                    _mark_combine_not_applicable(
                        document_id,
                        "COMBINE_NOT_APPLICABLE",
                        "Unhandled document exception prevented archive/combine.",
                    )
                    trace_store.record_failure_event(
                        document_id,
                        "DOCUMENT_PROCESSING_EXCEPTION",
                        "pipeline",
                        err_msg,
                        severity="failure",
                    )
                    import traceback
                    self._record_error_log(save_path, info.get('subject', file_name), f"代码执行异常: {err_msg} - {traceback.format_exc().splitlines()[-1] if traceback.format_exc().splitlines() else ''}")
                    processed_filepaths.add(pdf_path) # [Fix] Always mark as processed when exception hits
                finally:
                    time.sleep(0.5)

            # --- PHASE 1: EXECUTE RECONCILIATION ---
            import shutil
            orphans = []
            for og_path, og_info in ground_truth_files.items():
                if og_path not in processed_filepaths and os.path.exists(og_path):
                    orphans.append((og_path, og_info))
            
            if orphans:
                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "对账:", "color": "text-red-500", "msg": f"警告：发现 {len(orphans)} 个未被处理的遗漏附件，正在强制转移至待处理区 (raw_invoices)。"})
                raw_invoices_dir = os.path.join(save_path, "raw_invoices")
                os.makedirs(raw_invoices_dir, exist_ok=True)
                
                import uuid
                for orphan_path, orphan_info in orphans:
                    try:
                        orphan_filename = os.path.basename(orphan_path)
                        short_uuid = uuid.uuid4().hex[:6]
                        final_orphan_path = os.path.join(raw_invoices_dir, f"遗漏_{short_uuid}_{orphan_filename}")
                        shutil.copy2(orphan_path, final_orphan_path)
                        self.audit_counts["raw_invoices"] += 1
                        orphan_document_id = None
                        for record in trace_store.iter_records():
                            if record.get("source_path") == orphan_path:
                                orphan_document_id = record["document_id"]
                                break
                        
                        self.error_invoices.append({
                            "id": f"inv_orphan_{time.time()}_{len(self.error_invoices)}",
                            "date": "---",
                            "amount": "---",
                            "category": "系统遗漏",
                            "merchant": "未知开票方",
                            "path": final_orphan_path,
                            "name": orphan_filename,
                            "sColor": "bg-red-500",
                            "status": "处理中断遗漏",
                            "reason": "Pipeline 断层遗漏",
                            "rColor": "text-red-600 border-red-200 bg-red-50"
                        })
                        self.stats["errors"] += 1
                        self._record_error_log(save_path, orphan_info.get('subject', orphan_filename), "Pipeline 断层对账遗漏兜底")
                        self._safe_emit_artifact_event(
                            "raw_invoices",
                            final_orphan_path,
                            document_id=orphan_document_id,
                            source_kind=orphan_info.get("source_kind"),
                            reason_code="PIPELINE_ORPHAN_RECOVERED",
                            category="raw_invoices",
                            extra=self._attachment_diag_metadata(
                                orphan_info,
                                file_name=orphan_filename,
                                document_id=orphan_document_id,
                            ),
                        )
                        if orphan_document_id:
                            trace_store.set_fields(
                                orphan_document_id,
                                naming_result={"status": "skipped", "reason_code": "PIPELINE_ORPHAN_RECOVERED"},
                                archive_target=final_orphan_path,
                            )
                            _mark_combine_not_applicable(
                                orphan_document_id,
                                "COMBINE_NOT_APPLICABLE",
                                "Orphaned document was recovered after pipeline reconciliation.",
                            )
                            trace_store.record_failure_event(
                                orphan_document_id,
                                "PIPELINE_ORPHAN_RECOVERED",
                                "reconciliation",
                                "Pipeline 断层遗漏，已进入 raw_invoices 兜底。",
                                severity="failure",
                            )
                    except Exception as e:
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"无法转移遗漏文件 {orphan_path}: {str(e)}"})
                        
            # --- PHASE 2: 票据撮合与交替重命名 ---
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-blue-400", "msg": "开始执行发票↔行程单/水单撮合..."})
            
            def _parse_archived_filename(filename):
                """从归档文件名解析元数据: 日期_类型_金额_商户.ext 或 日期-出发-到达-火车票.ext"""
                import re
                name, ext = os.path.splitext(filename)
                parts = name.split("_")
                if len(parts) >= 4:
                    return {"date": parts[0], "type": parts[1], "amount": parts[2], "seller": "_".join(parts[3:]), "ext": ext}
                return {"date": parts[0] if parts else "", "type": "", "amount": "", "seller": "", "ext": ext}
            
            def _reconcile_ride_documents(ride_folder_path):
                """打车文件夹内的发票↔行程单撮合 + 交替重命名"""
                if not os.path.isdir(ride_folder_path):
                    return
                
                invoices = []
                itineraries = []
                
                for fname in os.listdir(ride_folder_path):
                    fpath = os.path.join(ride_folder_path, fname)
                    if not os.path.isfile(fpath):
                        continue
                    meta = _parse_archived_filename(fname)
                    meta["path"] = fpath
                    meta["filename"] = fname
                    meta["document_id"] = trace_store.get_document_id_by_archive_target(fpath)
                    
                    # 判断是发票还是行程单
                    is_itn = any(kw in fname for kw in ["行程单", "行程报销单", "报销单"])
                    if is_itn:
                        itineraries.append(meta)
                        _record_combine_candidate(meta["document_id"], "ride", "itinerary", meta)
                    else:
                        invoices.append(meta)
                        _record_combine_candidate(meta["document_id"], "ride", "invoice", meta)
                
                if not invoices or not itineraries:
                    reason_code = "RIDE_COMBINE_INSUFFICIENT_CANDIDATES"
                    for meta in invoices + itineraries:
                        _record_combine_result(
                            meta.get("document_id"),
                            "not_matched",
                            reason_code,
                            "打车目录缺少发票或行程单，无法撮合。",
                        )
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-gray-400", "msg": f"打车目录: 发票{len(invoices)}张, 行程单{len(itineraries)}张, 无法撮合"})
                    return
                
                # 联合主键匹配 (日期 + 金额, 含 1.03 税务容差)
                matched = []
                used_itn = set()
                
                for inv in invoices:
                    try:
                        inv_amt = float(inv["amount"])
                    except (ValueError, TypeError):
                        continue
                    inv_date = inv["date"]
                    
                    for j, itn in enumerate(itineraries):
                        if j in used_itn:
                            continue
                        try:
                            itn_amt = float(itn["amount"])
                        except (ValueError, TypeError):
                            continue
                        itn_date = itn["date"]
                        
                        # 金额匹配: 精确匹配 或 发票*1.03 ≈ 行程单 (税务容差)
                        # 注意: 打车发票是月末集中开具, 日期与行程单不同, 不做日期约束
                        if (abs(inv_amt - itn_amt) < 0.01 or 
                            abs(inv_amt * 1.03 - itn_amt) < 0.50 or
                            abs(itn_amt * 1.03 - inv_amt) < 0.50):
                            matched.append((inv, itn))
                            used_itn.add(j)
                            break
                
                # 交替重命名
                for idx, (inv, itn) in enumerate(matched, 1):
                    inv_document_id = inv.get("document_id")
                    itn_document_id = itn.get("document_id")
                    try:
                        mmdd = inv["date"][4:8] if len(inv["date"]) >= 8 else inv["date"]
                        # 以行程单含税金额为基准
                        try:
                            base_amount = f"{float(itn['amount']):.2f}"
                        except (ValueError, TypeError):
                            base_amount = itn["amount"]
                        
                        # 判断平台 (滴滴 vs 高德)
                        platform = "滴滴"
                        if any(kw in inv["filename"] for kw in ["高德", "约车", "盛智"]):
                            platform = "高德"
                        if any(kw in itn["filename"] for kw in ["高德", "约车"]):
                            platform = "高德"
                        
                        inv_ext = inv["ext"]
                        itn_ext = itn["ext"]
                        inv_new = f"{mmdd}-{platform}-{idx:02d}-发票_{base_amount}元{inv_ext}"
                        itn_new = f"{mmdd}-{platform}-{idx:02d}-行程单_{base_amount}元{itn_ext}"
                        
                        inv_new_path = os.path.join(ride_folder_path, inv_new)
                        itn_new_path = os.path.join(ride_folder_path, itn_new)
                        
                        os.rename(inv["path"], inv_new_path)
                        os.rename(itn["path"], itn_new_path)
                        trace_store.move_archive_target(inv["path"], inv_new_path)
                        trace_store.move_archive_target(itn["path"], itn_new_path)
                        _record_combine_result(
                            inv_document_id,
                            "matched",
                            "RIDE_COMBINE_MATCHED",
                            "打车发票已成功撮合并重命名。",
                            paired_with=os.path.basename(itn_new_path),
                            pair_index=idx,
                            final_filename=inv_new,
                        )
                        _record_combine_result(
                            itn_document_id,
                            "matched",
                            "RIDE_COMBINE_MATCHED",
                            "打车行程单已成功撮合并重命名。",
                            paired_with=os.path.basename(inv_new_path),
                            pair_index=idx,
                            final_filename=itn_new,
                        )
                        
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-blue-400", "msg": f"✅ {platform}配对 #{idx}: {inv_new} ↔ {itn_new}"})
                    except Exception as e:
                        _record_combine_result(
                            inv_document_id,
                            "rename_failed",
                            "RIDE_COMBINE_RENAME_FAILED",
                            str(e),
                        )
                        _record_combine_result(
                            itn_document_id,
                            "rename_failed",
                            "RIDE_COMBINE_RENAME_FAILED",
                            str(e),
                        )
                        trace_store.record_failure_event(
                            inv_document_id,
                            "RIDE_COMBINE_RENAME_FAILED",
                            "combine",
                            str(e),
                            severity="failure",
                        )
                        trace_store.record_failure_event(
                            itn_document_id,
                            "RIDE_COMBINE_RENAME_FAILED",
                            "combine",
                            str(e),
                            severity="failure",
                        )
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"重命名失败: {e}"})
                
                # 报告未匹配
                unmatched_inv = [inv for inv in invoices if inv not in [m[0] for m in matched]]
                unmatched_itn = [itn for itn in itineraries if itn not in [m[1] for m in matched]]
                for inv in unmatched_inv:
                    _record_combine_result(
                        inv.get("document_id"),
                        "not_matched",
                        "RIDE_COMBINE_NO_MATCH",
                        "未找到可匹配的打车行程单。",
                    )
                for itn in unmatched_itn:
                    _record_combine_result(
                        itn.get("document_id"),
                        "not_matched",
                        "RIDE_COMBINE_NO_MATCH",
                        "未找到可匹配的打车发票。",
                    )
                if unmatched_inv:
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-yellow-400", "msg": f"未匹配发票 {len(unmatched_inv)} 张: {', '.join(u['filename'] for u in unmatched_inv)}"})
                if unmatched_itn:
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-yellow-400", "msg": f"未匹配行程单 {len(unmatched_itn)} 张: {', '.join(u['filename'] for u in unmatched_itn)}"})
                
                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-blue-400", "msg": f"打车撮合完成: 成功配对 {len(matched)} 组"})
            
            def _reconcile_hotel_documents(hotel_folder_path):
                """住宿文件夹内的发票↔水单撮合 + 交替重命名"""
                if not os.path.isdir(hotel_folder_path):
                    return
                
                invoices = []
                folios = []
                
                for fname in os.listdir(hotel_folder_path):
                    fpath = os.path.join(hotel_folder_path, fname)
                    if not os.path.isfile(fpath):
                        continue
                    meta = _parse_archived_filename(fname)
                    meta["path"] = fpath
                    meta["filename"] = fname
                    meta["document_id"] = trace_store.get_document_id_by_archive_target(fpath)
                    
                    # 住宿确认单/航班行程单不参与发票↔水单撮合
                    is_order = any(kw in fname for kw in ["确认单", "行程单"])
                    if is_order:
                        _record_combine_result(
                            meta.get("document_id"),
                            "not_applicable",
                            "COMBINE_ORDER_EXCLUDED",
                            "订单确认单不参与发票↔水单撮合。",
                        )
                        continue
                    is_folio = any(kw in fname.lower() for kw in ["水单", "folio", "账单", "明细"])
                    if is_folio:
                        folios.append(meta)
                        _record_combine_candidate(meta["document_id"], "hotel", "folio", meta)
                    else:
                        invoices.append(meta)
                        _record_combine_candidate(meta["document_id"], "hotel", "invoice", meta)
                
                if not invoices or not folios:
                    reason_code = "HOTEL_COMBINE_INSUFFICIENT_CANDIDATES"
                    for meta in invoices + folios:
                        _record_combine_result(
                            meta.get("document_id"),
                            "not_matched",
                            reason_code,
                            "住宿目录缺少发票或水单，无法撮合。",
                        )
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-gray-400", "msg": f"住宿目录: 发票{len(invoices)}张, 水单{len(folios)}张, 无法撮合"})
                    return
                
                # 联合主键匹配 (金额精确 + 日期 0-3 天容差)
                from datetime import datetime as dt, timedelta
                matched = []
                used_fol = set()
                
                for inv in invoices:
                    try:
                        inv_amt = float(inv["amount"])
                    except (ValueError, TypeError):
                        continue
                    
                    for j, fol in enumerate(folios):
                        if j in used_fol:
                            continue
                        try:
                            fol_amt = float(fol["amount"])
                        except (ValueError, TypeError):
                            continue
                        
                        # 金额必须精确匹配
                        if abs(inv_amt - fol_amt) > 0.01:
                            continue
                        
                        # 日期容差 0-3 天
                        try:
                            inv_d = dt.strptime(inv["date"], "%Y%m%d").date()
                            fol_d = dt.strptime(fol["date"], "%Y%m%d").date()
                            if abs((inv_d - fol_d).days) <= 3:
                                matched.append((inv, fol))
                                used_fol.add(j)
                                break
                        except (ValueError, TypeError):
                            # 日期解析失败则跳过日期检查，仅凭金额
                            matched.append((inv, fol))
                            used_fol.add(j)
                            break
                
                # 交替重命名
                for idx, (inv, fol) in enumerate(matched, 1):
                    inv_document_id = inv.get("document_id")
                    fol_document_id = fol.get("document_id")
                    try:
                        base_date = inv["date"]  # 以发票日期为基准
                        try:
                            base_amount = f"{float(inv['amount']):.2f}"
                        except (ValueError, TypeError):
                            base_amount = inv["amount"]
                        
                        inv_ext = inv["ext"]
                        fol_ext = fol["ext"]
                        inv_new = f"{base_date}-住宿-{idx:02d}-发票_{base_amount}元{inv_ext}"
                        fol_new = f"{base_date}-住宿-{idx:02d}-水单_{base_amount}元{fol_ext}"
                        
                        inv_new_path = os.path.join(hotel_folder_path, inv_new)
                        fol_new_path = os.path.join(hotel_folder_path, fol_new)
                        
                        os.rename(inv["path"], inv_new_path)
                        os.rename(fol["path"], fol_new_path)
                        trace_store.move_archive_target(inv["path"], inv_new_path)
                        trace_store.move_archive_target(fol["path"], fol_new_path)
                        _record_combine_result(
                            inv_document_id,
                            "matched",
                            "HOTEL_COMBINE_MATCHED",
                            "住宿发票已成功撮合并重命名。",
                            paired_with=os.path.basename(fol_new_path),
                            pair_index=idx,
                            final_filename=inv_new,
                        )
                        _record_combine_result(
                            fol_document_id,
                            "matched",
                            "HOTEL_COMBINE_MATCHED",
                            "住宿水单已成功撮合并重命名。",
                            paired_with=os.path.basename(inv_new_path),
                            pair_index=idx,
                            final_filename=fol_new,
                        )
                        
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-blue-400", "msg": f"✅ 住宿配对 #{idx}: {inv_new} ↔ {fol_new}"})
                    except Exception as e:
                        _record_combine_result(
                            inv_document_id,
                            "rename_failed",
                            "HOTEL_COMBINE_RENAME_FAILED",
                            str(e),
                        )
                        _record_combine_result(
                            fol_document_id,
                            "rename_failed",
                            "HOTEL_COMBINE_RENAME_FAILED",
                            str(e),
                        )
                        trace_store.record_failure_event(
                            inv_document_id,
                            "HOTEL_COMBINE_RENAME_FAILED",
                            "combine",
                            str(e),
                            severity="failure",
                        )
                        trace_store.record_failure_event(
                            fol_document_id,
                            "HOTEL_COMBINE_RENAME_FAILED",
                            "combine",
                            str(e),
                            severity="failure",
                        )
                        self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"重命名失败: {e}"})
                
                unmatched_inv = [inv for inv in invoices if inv not in [m[0] for m in matched]]
                unmatched_fol = [fol for fol in folios if fol not in [m[1] for m in matched]]
                for inv in unmatched_inv:
                    _record_combine_result(
                        inv.get("document_id"),
                        "not_matched",
                        "HOTEL_COMBINE_NO_MATCH",
                        "未找到可匹配的住宿水单。",
                    )
                for fol in unmatched_fol:
                    _record_combine_result(
                        fol.get("document_id"),
                        "not_matched",
                        "HOTEL_COMBINE_NO_MATCH",
                        "未找到可匹配的住宿发票。",
                    )

                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-blue-400", "msg": f"住宿撮合完成: 成功配对 {len(matched)} 组"})
            
            # 执行撮合
            ride_folder = os.path.join(save_path, "打车")
            hotel_folder = os.path.join(save_path, "住宿发票")
            
            try:
                _reconcile_ride_documents(ride_folder)
            except Exception as e:
                phase2_had_error = True
                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"打车撮合异常: {e}"})
            
            try:
                _reconcile_hotel_documents(hotel_folder)
            except Exception as e:
                phase2_had_error = True
                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"住宿撮合异常: {e}"})
            
            phase2_completed = True
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "完成:", "color": "text-emerald-400", "msg": "Phase 2 撮合全部完成"})

        except Exception as main_loop_err:
            loop_result = "failed"
            print(f">>> [错误] 核心处理循环发生异常: {main_loop_err}")
            self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"批量处理异常: {main_loop_err}"})
            self._safe_emit_stage_event("_run_processing_loop", "exit", {"result": "failed", "reason": str(main_loop_err)})
        finally:
            _finalize_trace_defaults()
            try:
                trace_store.flush()
            except Exception as trace_err:
                self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "错误:", "color": "text-red-400", "msg": f"诊断 trace 写入失败: {trace_err}"})
            if loop_result != "failed":
                self._safe_emit_stage_event(
                    "_run_processing_loop",
                    "exit",
                    {
                        "result": loop_result,
                        "attachments": total_attachments,
                    },
                )

    def _retain_artifact(self, save_path, source_path, bucket, reason, metadata=None):
        """在 staging 清理前保留一份可追踪的原件副本。"""
        import json
        import re
        import shutil
        import uuid

        retention_dir = os.path.join(save_path, "_audit_retention", bucket)
        os.makedirs(retention_dir, exist_ok=True)

        is_url_placeholder = (
            metadata
            and (
                metadata.get("source_kind") == "url"
                or str(source_path).startswith(("http://", "https://"))
            )
            and (not source_path or not os.path.exists(source_path))
        )

        if is_url_placeholder:
            subject = str((metadata or {}).get("subject", "LinkRetention"))
            safe_subject = re.sub(r"\s+", "_", re.sub(r'[\\/:*?"<>|]+', "_", subject)).strip(" _")[:40] or "LinkRetention"
            candidate_index = int((metadata or {}).get("candidate_index", 1) or 1)
            original_name = f"LinkRetention_{safe_subject}_{candidate_index}.url.txt"
            target_name = original_name
            target_path = os.path.join(retention_dir, target_name)
            while os.path.exists(target_path):
                stem, ext = os.path.splitext(original_name)
                target_name = f"{stem}_{uuid.uuid4().hex[:6]}{ext}"
                target_path = os.path.join(retention_dir, target_name)
            with open(target_path, "w", encoding="utf-8") as fh:
                fh.write(str(source_path))
        else:
            if not source_path or not os.path.exists(source_path):
                return source_path

            original_name = os.path.basename(source_path)
            target_name = original_name
            target_path = os.path.join(retention_dir, target_name)
            while os.path.exists(target_path):
                stem, ext = os.path.splitext(original_name)
                target_name = f"{stem}_{uuid.uuid4().hex[:6]}{ext}"
                target_path = os.path.join(retention_dir, target_name)

            shutil.copy2(source_path, target_path)

        sidecar = f"{target_path}.json"
        payload = {
            "reason": reason,
            "original_path": str(source_path),
            "retained_path": target_path,
            "captured_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        if metadata:
            payload["metadata"] = metadata

        try:
            with open(sidecar, "w", encoding="utf-8") as fh:
                json.dump(payload, fh, ensure_ascii=False, indent=2)
        except Exception as exc:
            print(f"Failed to write retention sidecar {sidecar}: {exc}")

        self.audit_counts["retention"] = int(self.audit_counts.get("retention", 0) or 0) + 1
        self._safe_emit_artifact_event(
            "retention",
            target_path,
            document_id=(metadata or {}).get("document_id"),
            source_kind=(metadata or {}).get("source_kind"),
            reason_code=(metadata or {}).get("prefilter_reason_code") or bucket,
            category=bucket,
            extra={
                "bucket": bucket,
                "retention_reason": reason,
                "metadata": metadata or {},
            },
        )

        return target_path

    def _send_to_manual_check(self, save_path, source_path, reason, metadata=None, is_url=False):
        """把待人工复核候选写入用户输出目录下的 Manual_Check。"""
        import json
        import re
        import shutil
        import uuid

        manual_dir = os.path.join(save_path, "Manual_Check")
        os.makedirs(manual_dir, exist_ok=True)

        def _unique_path(filename):
            target_path = os.path.join(manual_dir, filename)
            while os.path.exists(target_path):
                stem, ext = os.path.splitext(filename)
                filename_local = f"{stem}_{uuid.uuid4().hex[:6]}{ext}"
                target_path = os.path.join(manual_dir, filename_local)
                filename = filename_local
            return target_path

        if is_url:
            subject = ""
            if metadata:
                subject = str(metadata.get("subject", "LinkReview"))
            safe_subject = re.sub(r'\s+', '_', re.sub(r'[\\/:*?"<>|]+', '_', subject)).strip(" _")[:40] or "LinkReview"
            candidate_index = 1
            if metadata:
                candidate_index = int(metadata.get("candidate_index", 1) or 1)
            target_path = _unique_path(f"P0_LinkReview_{safe_subject}_{candidate_index}.url.txt")
            with open(target_path, "w", encoding="utf-8") as fh:
                fh.write(str(source_path))
            original_path = str(source_path)
        else:
            if not source_path or not os.path.exists(source_path):
                return source_path
            original_name = os.path.basename(source_path)
            prefix = "P0_Review"
            if metadata and metadata.get("file_name"):
                original_name = os.path.basename(str(metadata["file_name"]))
            target_path = _unique_path(f"{prefix}_{original_name}")
            shutil.copy2(source_path, target_path)
            original_path = source_path

        sidecar = f"{target_path}.json"
        payload = {
            "reason": reason,
            "original_path": original_path,
            "review_path": target_path,
            "captured_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "is_url": is_url,
        }
        if metadata:
            payload["metadata"] = metadata

        try:
            with open(sidecar, "w", encoding="utf-8") as fh:
                json.dump(payload, fh, ensure_ascii=False, indent=2)
        except Exception as exc:
            print(f"Failed to write manual-check sidecar {sidecar}: {exc}")

        self.audit_counts["manual_check"] = int(self.audit_counts.get("manual_check", 0) or 0) + 1
        self._safe_emit_artifact_event(
            "manual_check",
            target_path,
            document_id=(metadata or {}).get("document_id"),
            source_kind=(metadata or {}).get("source_kind"),
            reason_code=reason,
            category="Manual_Check",
            extra={
                "is_url": bool(is_url),
                "metadata": metadata or {},
            },
        )

        return target_path

    def _cwt_cancellation_matching(self, save_path):
        """CWT 取消撮合: 根据取消知会中的人名，在已归档的住宿确认单中寻找匹配的预订，
        将匹配的预订也移入 Manual_Check，并在 sidecar 中注明撮合关系。"""
        import re
        import shutil
        import json

        cancellations = getattr(self, '_cwt_cancellation_registry', [])
        if not cancellations:
            return

        # 从取消知会文件名中提取人名
        # 格式: 酒店预定取消知会-{name}-{date}入住-{city} (CONNECT 订单号：{order}).pdf
        cancel_pattern = re.compile(r'取消知会[_\-]?(\S+?)[_\-]\d')

        manual_dir = os.path.join(save_path, "Manual_Check")
        hotel_dir = os.path.join(save_path, "住宿发票")

        for cancel in cancellations:
            cancel_fn = cancel.get("file_name", "")
            m = cancel_pattern.search(cancel_fn)
            if not m:
                continue
            person_name = m.group(1)
            if not person_name or len(person_name) < 2:
                continue

            # 搜索住宿发票目录中匹配此人名的住宿确认单
            if not os.path.isdir(hotel_dir):
                continue
            for fn in os.listdir(hotel_dir):
                if person_name in fn and "酒店" in fn.lower():
                    src = os.path.join(hotel_dir, fn)
                    dst = os.path.join(manual_dir, f"P0_CancelMatch_{fn}")
                    if os.path.exists(dst):
                        continue
                    os.makedirs(manual_dir, exist_ok=True)
                    shutil.move(src, dst)
                    # 写 sidecar
                    sidecar = {
                        "reason": "CWT_CANCELLATION_MATCH",
                        "matched_cancellation": cancel_fn,
                        "matched_person": person_name,
                        "original_path": src,
                        "review_path": dst,
                        "captured_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                    try:
                        with open(f"{dst}.json", "w", encoding="utf-8") as fh:
                            json.dump(sidecar, fh, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                    self.logs.append({"time": time.strftime("[%H:%M:%S]"), "type": "撮合:", "color": "text-amber-400", "msg": f"匹配到取消对应的预订: {fn} ↔ {cancel_fn}"})

    def _record_error_log(self, save_path, email_title, error_reason):
        """记录错误日志到对应的 csv 文件中"""
        import os
        import csv
        from datetime import datetime
        
        # 目录保护
        if not os.path.exists(save_path):
            os.makedirs(save_path, exist_ok=True)
            
        log_file = os.path.join(save_path, "异常发票处理日志.csv")
        file_exists = os.path.exists(log_file)
        
        try:
            with open(log_file, 'a', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ['处理时间', '邮件标题', '错误原因']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                if not file_exists:
                    writer.writeheader()
                writer.writerow({
                    '处理时间': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    '邮件标题': email_title,
                    '错误原因': error_reason
                })
        except Exception as e:
            print(f"Failed to write error log: {e}")

    def _cleanup_temp_folders(self):
        """自动清理程序执行过程中产生的临时文件夹"""
        import os
        import shutil
        current_dir = os.getcwd()
        if self._run_context.get("enabled"):
            temp_paths = [
                self._run_context.get("staging_dir"),
                os.path.join(self._run_context.get("run_root", ""), "temp"),
            ]
        else:
            temp_paths = [os.path.join(current_dir, t_dir) for t_dir in ["staging", "temp"]]

        for target_path in temp_paths:
            if not target_path or not os.path.exists(target_path) or not os.path.isdir(target_path):
                continue
            try:
                shutil.rmtree(target_path)
                print(f"Cleaned up temp folder: {target_path}")
            except Exception as e:
                print(f"Failed to clean up {target_path}: {e}")
                self.logs.append({
                    "time": time.strftime("[%H:%M:%S]"),
                    "type": "ERROR",
                    "color": "text-error",
                    "msg": f"Failed to clean temporary folder {target_path}: {e}",
                })

    def _legacy_export_result_detail_pre_release_prep(self, export_path=""):
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"success": False, "message": "缺少 openpyxl 依赖，无法导出 Excel"}

        target_dir = export_path or self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        os.makedirs(target_dir, exist_ok=True)

        def _parse_amount(value):
            text = str(value or "").strip()
            if not text:
                return 0.0
            cleaned = re.sub(r"[^\d.\-]", "", text)
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except Exception:
                return 0.0

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "分类汇总"
        summary_sheet.append(["发票分类", "数量", "金额合计"])

        category_summary = {}
        for item in self.processed_invoices:
            category = str(item.get("category") or "未分类")
            bucket = category_summary.setdefault(category, {"count": 0, "amount": 0.0})
            bucket["count"] += 1
            bucket["amount"] += _parse_amount(item.get("amount", ""))

        if category_summary:
            for category in sorted(category_summary.keys()):
                bucket = category_summary[category]
                summary_sheet.append([category, bucket["count"], round(bucket["amount"], 2)])
        else:
            summary_sheet.append(["暂无成功记录", 0, 0.0])

        success_sheet = workbook.create_sheet("成功明细")
        success_sheet.append(["日期", "金额", "销售方", "分类", "文件路径"])
        for item in self.processed_invoices:
            success_sheet.append([
                item.get("date", ""),
                item.get("amount", ""),
                item.get("merchant", ""),
                item.get("category", ""),
                item.get("path", ""),
            ])

        error_sheet = workbook.create_sheet("异常记录")
        error_sheet.append(["分组", "状态", "原因", "日期", "金额", "销售方", "文件名", "文件路径"])
        for group in self._group_error_invoices():
            for item in group.get("items", []):
                error_sheet.append([
                    group.get("label", ""),
                    item.get("status", ""),
                    item.get("reason", ""),
                    item.get("date", ""),
                    item.get("amount", ""),
                    item.get("merchant", ""),
                    item.get("name", ""),
                    item.get("path", ""),
                ])

        export_file = os.path.join(target_dir, f"结果明细_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        workbook.save(export_file)
        self._last_export_path = export_file
        return {"success": True, "message": "结果明细已导出", "path": export_file}

    def _legacy_start_processing_pre_release_prep(self, rules_text, save_path, date_from=None, date_to=None, email_address=None, auth_code=None, api_key=None):
        '''
        """前端点击执行时调用：新开线程防止卡死界面"""
        import os
        print(f"Start processing with rules: {rules_text}, save to: {save_path}, date range: {date_from} to {date_to}")
        if self._is_running or (self._worker_thread and self._worker_thread.is_alive()):
            return {"success": False, "message": "任务已在运行中"}
            
        # 目录保护：启动前先确保保存路径存在
        try:
            if not os.path.exists(save_path):
                os.makedirs(save_path, exist_ok=True)
        except Exception as e:
            return {"success": False, "message": f"无法创建保存目录: {e}"}
            
        self._set_run_state("running", status_text="鍑嗗涓?", progress=5, last_error="")
        self.status_text = "准备中"
        
        thread = threading.Thread(target=self._processing_worker, args=(rules_text, save_path, date_from, date_to, email_address, auth_code, api_key))
        thread.daemon = True
        thread.start()
        self._worker_thread = thread
        return {"success": True, "message": "任务已启动"}

        '''

    def _start_async_finalizers(self, fetcher=None):
        def _runner():
            self._safe_emit_stage_event("cleanup_finalize", "enter")
            self._cleanup_temp_folders()

            if fetcher is not None:
                try:
                    fetcher.disconnect()
                except Exception as exc:
                    self.logs.append({
                        "time": time.strftime("[%H:%M:%S]"),
                        "type": "ERROR",
                        "color": "text-error",
                        "msg": f"Failed to disconnect mailbox cleanly: {exc}",
                    })
            self._safe_emit_stage_event("cleanup_finalize", "exit")

        cleanup_thread = threading.Thread(target=_runner, daemon=True)
        cleanup_thread.start()
        return cleanup_thread

    def _mark_finalizing(self):
        finalizing_progress = self.progress if self.progress >= 95 else 99
        self._set_run_state("finalizing", status_text="正在收尾...", progress=finalizing_progress)

    def _fail_run(self, status_text, error_message, fetcher=None, include_traceback=False):
        active_exc = sys.exc_info()[1]
        self._packaged_diag_write(
            "fail_run",
            "_fail_run",
            "exception" if active_exc is not None else "failure",
            summary={"include_traceback": bool(include_traceback)},
            exc=active_exc,
        )
        self._mark_finalizing()
        self._start_async_finalizers(fetcher)

        if include_traceback:
            import traceback
            error_message = f"{error_message} | {traceback.format_exc()}"

        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "ERROR",
            "color": "text-error",
            "msg": f"System exception: {error_message}",
        })
        self._finish_run(False, status_text, last_error=error_message)

    def _validate_date_range(self, date_from, date_to):
        date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
        if date_from and not date_pattern.match(str(date_from)):
            return "开始日期格式必须为 YYYY-MM-DD"
        if date_to and not date_pattern.match(str(date_to)):
            return "结束日期格式必须为 YYYY-MM-DD"
        if date_from and date_to and date_from > date_to:
            return "开始日期不能晚于结束日期"
        return ""

    def _manual_check_path(self):
        base_path = self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        return os.path.join(base_path, "Manual_Check")

    def _build_error_breakdown(self):
        breakdown = {
            "manual_review": 0,
            "retained_record": 0,
            "processing_error": 0,
        }
        reason_codes = {}

        for item in list(self.error_invoices):
            group_key, _group_label = self._classify_error_invoice(item)
            if group_key not in breakdown:
                group_key = "processing_error"
            breakdown[group_key] += 1
            reason_code = str((item or {}).get("reason", "") or "").strip()
            if reason_code:
                reason_codes[reason_code] = int(reason_codes.get(reason_code, 0) or 0) + 1

        breakdown["reason_codes"] = reason_codes
        return breakdown

    def _summarize_stats(self):
        breakdown = self._build_error_breakdown()
        return {
            "emails": int(self.stats.get("emails", 0) or 0),
            "success_count": len(self.processed_invoices),
            "error_count": len(self.error_invoices),
            "manual_check_count": int(breakdown.get("manual_review", 0) or self.audit_counts.get("manual_check", 0) or 0),
            "retention_count": int(breakdown.get("retained_record", 0) or 0),
            "raw_invoice_count": int(self.audit_counts.get("raw_invoices", 0) or 0),
            "processing_error_count": int(breakdown.get("processing_error", 0) or 0),
            "result_breakdown": {
                "manual_review": int(breakdown.get("manual_review", 0) or 0),
                "retained_record": int(breakdown.get("retained_record", 0) or 0),
                "processing_error": int(breakdown.get("processing_error", 0) or 0),
            },
            "reason_code_breakdown": breakdown.get("reason_codes", {}),
            "quota_exhausted": bool(self.quota_exhausted),
            "quota_message": self.quota_message,
            "run_state": self.run_state,
            "status_text": self.status_text,
        }

    def _classify_error_invoice(self, item):
        path = str((item or {}).get("path", "") or "")
        status = str((item or {}).get("status", "") or "")
        reason = str((item or {}).get("reason", "") or "")
        category = str((item or {}).get("category", "") or "")
        merchant = str((item or {}).get("merchant", "") or "")
        normalized = " ".join([path, status, reason, category, merchant]).lower()

        if "manual_check" in normalized or "人工复核" in normalized or "待人工复核" in normalized:
            return "manual_review", "待人工复核"
        if any(token in normalized for token in [
            "_audit_retention",
            "controlled_run_non_provider_url",
            "history_skipped",
            "prefilter_b_retained",
            "provider_recovery_failed",
            "url_non_invoice_page_skipped",
            "model_rejected",
            "受控跑批保全",
            "已保全待判断",
            "保全",
            "保留",
            "retention",
        ]):
            return "retained_record", "保留记录"
        if any(token in normalized for token in [
            "url_page_timeout",
            "url_auth_wall_detected",
            "url_no_response",
            "url_download_failed",
            "链接下载失败",
        ]):
            return "processing_error", "真实处理异常"
        if self._resolve_quota_message(normalized):
            return "processing_error", "真实处理异常"
        if any(token in normalized for token in [
            "processing_errors",
            "处理中断遗漏",
            "pipeline 断层遗漏",
            "系统遗漏",
            "处理单张票据时抛出异常",
            "异常",
            "failed",
            "error",
        ]):
            return "processing_error", "真实处理异常"
        return "processing_error", "真实处理异常"

    def _group_error_invoices(self):
        grouped = {}
        for item in list(self.error_invoices):
            group_key, group_label = self._classify_error_invoice(item)
            group = grouped.setdefault(group_key, {"key": group_key, "label": group_label, "count": 0, "items": []})
            enriched = dict(item)
            enriched["groupKey"] = group_key
            enriched["groupLabel"] = group_label
            group["items"].append(enriched)
            group["count"] += 1
        return list(grouped.values())

    def stop_processing(self):
        if not self._is_running or not (self._worker_thread and self._worker_thread.is_alive()):
            return {"success": False, "message": "当前没有正在运行的任务"}
        self._request_safe_stop()
        return {"success": True, "message": "已收到停止指令，当前文件处理完成后将结束任务"}

    def close_window(self):
        try:
            import webview
        except ImportError:
            return {"success": False, "message": "桌面窗口接口不可用"}

        if not webview.windows:
            return {"success": False, "message": "当前没有可关闭的桌面窗口"}

        window = webview.windows[0]

        def _shutdown():
            try:
                window.destroy()
            finally:
                time.sleep(0.15)
                os._exit(0)

        threading.Thread(target=_shutdown, daemon=True).start()
        return {"success": True, "message": "窗口关闭中"}

    def minimize_window(self):
        try:
            import webview
        except ImportError:
            return {"success": False, "message": "桌面窗口接口不可用"}

        if not webview.windows:
            return {"success": False, "message": "当前没有可最小化的桌面窗口"}

        window = webview.windows[0]

        try:
            window.minimize()
            return {"success": True, "message": "窗口已最小化"}
        except Exception as exc:
            return {"success": False, "message": f"窗口最小化失败: {exc}"}

    def open_manual_check_folder(self):
        manual_path = self._manual_check_path()
        return self.open_folder(manual_path)

    def _legacy_export_run_summary_pre_release_prep(self, export_path=""):
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"success": False, "message": "缺少 openpyxl 依赖，无法导出 Excel"}

        target_dir = export_path or self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        os.makedirs(target_dir, exist_ok=True)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "运行汇总"
        summary_sheet.append(["项目", "值"])
        for key, value in self._summarize_stats().items():
            summary_sheet.append([key, value])

        success_sheet = workbook.create_sheet("成功记录")
        success_sheet.append(["日期", "金额", "销售方", "分类", "文件路径"])
        for item in self.processed_invoices:
            success_sheet.append([
                item.get("date", ""),
                item.get("amount", ""),
                item.get("merchant", ""),
                item.get("category", ""),
                item.get("path", ""),
            ])

        error_sheet = workbook.create_sheet("异常记录")
        error_sheet.append(["分组", "状态", "原因", "日期", "金额", "销售方", "文件名", "文件路径"])
        for group in self._group_error_invoices():
            for item in group.get("items", []):
                error_sheet.append([
                    group.get("label", ""),
                    item.get("status", ""),
                    item.get("reason", ""),
                    item.get("date", ""),
                    item.get("amount", ""),
                    item.get("merchant", ""),
                    item.get("name", ""),
                    item.get("path", ""),
                ])

        export_file = os.path.join(target_dir, f"运行摘要_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        workbook.save(export_file)
        self._last_export_path = export_file
        return {"success": True, "message": "运行摘要已导出", "path": export_file}

    def start_processing(self, rules_text, save_path, date_from=None, date_to=None, email_address=None, auth_code=None, api_key=None):
        import os

        if self._is_running or (self._worker_thread and self._worker_thread.is_alive()):
            return {"success": False, "message": "任务已在运行中"}

        if not email_address or not auth_code or not api_key:
            return {"success": False, "message": "缺少必要凭证，请填写邮箱、授权码和 API Key"}

        run_context = self._refresh_run_context()
        requested_save_path = save_path or self.get_default_save_path()
        effective_save_path = self._effective_save_dir(requested_save_path)
        effective_date_from, effective_date_to = self._effective_date_range(date_from, date_to)
        date_error = self._validate_date_range(effective_date_from, effective_date_to)
        if date_error:
            return {"success": False, "message": date_error}

        print(
            "Start processing",
            {
                "run_context_enabled": bool(run_context.get("enabled")),
                "run_id": run_context.get("run_id", ""),
                "requested_save_path": requested_save_path,
                "effective_save_path": effective_save_path,
                "requested_date_from": date_from,
                "requested_date_to": date_to,
                "effective_date_from": effective_date_from,
                "effective_date_to": effective_date_to,
            },
        )
        self._requested_save_path = requested_save_path
        self._effective_save_path = effective_save_path
        self._effective_date_from = effective_date_from or ""
        self._effective_date_to = effective_date_to or ""
        self._current_run_id = run_context.get("run_id", "")
        ensure_run_context_dirs(run_context)

        try:
            if effective_save_path and not os.path.exists(effective_save_path):
                os.makedirs(effective_save_path, exist_ok=True)
        except Exception as exc:
            return {"success": False, "message": f"无法创建输出目录: {exc}"}

        remember_settings = bool((self._settings_store.load() or {}).get("remember_settings", True))
        if remember_settings:
            self.save_user_settings({
                "email": email_address,
                "auth_code": auth_code,
                "api_key": api_key,
                "save_path": effective_save_path,
                "date_from": effective_date_from or "",
                "date_to": effective_date_to or "",
                "remember_settings": True,
            })
        else:
            self.save_user_settings({"remember_settings": False})

        self._packaged_diag_reset(
            {
                "requested_save_path": requested_save_path,
                "effective_save_path": effective_save_path,
                "date_from": self._effective_date_from,
                "date_to": self._effective_date_to,
                "email_domain": self._packaged_diag_email_domain(email_address),
                "has_auth_code": bool(auth_code),
                "has_api_key": bool(api_key),
            }
        )
        self._set_run_state("running", status_text="正在准备运行...", progress=5, last_error="")
        self._packaged_diag_write(
            "progress_5_written",
            "start_processing",
            "success",
            summary={
                "requested_save_path": requested_save_path,
                "effective_save_path": effective_save_path,
                "date_from": self._effective_date_from,
                "date_to": self._effective_date_to,
                "email_domain": self._packaged_diag_email_domain(email_address),
                "has_auth_code": bool(auth_code),
                "has_api_key": bool(api_key),
            },
        )
        self._safe_emit_stage_event(
            "start_processing",
            "enter",
            {
                "requested_save_path": requested_save_path,
                "effective_save_path": effective_save_path,
                "date_from": self._effective_date_from,
                "date_to": self._effective_date_to,
                **self._sensitive_summary(email_address, auth_code, api_key),
            },
        )
        self.logs.append({
            "time": time.strftime("[%H:%M:%S]"),
            "type": "信息",
            "color": "text-blue-400",
            "msg": "前端请求已接收，后台任务正在启动。",
        })
        self._safe_write_run_config(email_address, auth_code=auth_code, api_key=api_key)
        self._start_truth_audit_async(email_address, auth_code)

        thread = threading.Thread(
            target=self._processing_worker,
            args=(rules_text, effective_save_path, self._effective_date_from, self._effective_date_to, email_address, auth_code, api_key),
            name="InvoiceFlowWorker",
            daemon=True,
        )
        self._packaged_diag_write("worker_thread_created", "start_processing", "success")
        self._worker_thread = thread
        thread.start()
        self._packaged_diag_write(
            "worker_thread_started",
            "start_processing",
            "success",
            summary={"thread_is_alive": bool(thread.is_alive())},
        )
        self._safe_emit_stage_event("start_processing", "exit", {"result": "started"})
        return {"success": True, "message": "任务已启动"}

    def get_processed_records(self):
        """前端数据分析页面调用，获取已处理的账单或发票记录"""
        return self.processed_invoices

    def get_progress(self):
        """前端轮询进度条和日志调用"""
        if not self._is_running and self.progress == 0:
            payload = {
                "progress": 0,
                "status_text": "等待任务开始...",
                "logs": [],
                "new_categories": [],
                "stats": {"emails": 0, "invoices": 0, "errors": 0},
                "is_running": False,
                "run_state": self.run_state,
                "last_error": self.last_error,
                "stop_requested": self._stop_requested,
                "can_stop": False,
                "quota_exhausted": self.quota_exhausted,
                "quota_message": self.quota_message,
            }
        else:
            payload = {
                "progress": self.progress,
                "status_text": self.status_text,
                "logs": self.logs[-20:], # 返回最新的日志
                "new_categories": list(self.discovered_categories), # 传回后端发现的所有分类
                "stats": getattr(self, "stats", {"emails": 0, "invoices": 0, "errors": 0}),
                "is_running": self._is_running,
                "run_state": self.run_state,
                "last_error": self.last_error,
                "stop_requested": self._stop_requested,
                "can_stop": bool(self._is_running and not self._stop_requested),
                "quota_exhausted": self.quota_exhausted,
                "quota_message": self.quota_message,
            }
        self._packaged_diag_log_progress_poll(payload)
        return payload

    def get_results(self):
        """前端分析页调用，获取最终的统计数据"""
        grouped_errors = self._group_error_invoices()
        summary = self._summarize_stats()
        return {
            "categories": list(self.discovered_categories),
            "successInvoices": self.processed_invoices,
            "errorInvoices": self.error_invoices,
            "groupedErrorInvoices": grouped_errors,
            "manual_check_path": self._manual_check_path(),
            "output_path": self._effective_save_path or self._requested_save_path or self.get_default_save_path(),
            "summary": summary,
            "resultBreakdown": summary.get("result_breakdown", {}),
            "reasonCodeBreakdown": summary.get("reason_code_breakdown", {}),
            "quota_exhausted": self.quota_exhausted,
            "quota_message": self.quota_message,
            "last_export_path": self._last_export_path,
            "invoices": self.processed_invoices # 兼容旧的数据结构
        }

    def choose_directory(self):
        """调用系统原生目录选择器"""
        run_context = self._refresh_run_context()
        if run_context.get("enabled"):
            return {"success": True, "path": run_context.get("output_dir", "")}
        print("Opening directory dialog...")
        import webview
        # webview.windows 列表包含了所有当前活动的窗口，取第一个
        if webview.windows:
            window = webview.windows[0]
            # webview.FOLDER_DIALOG
            result = window.create_file_dialog(webview.FOLDER_DIALOG)
            if result and len(result) > 0:
                print(f"Selected: {result[0]}")
                return {"success": True, "path": result[0]}
        return {"success": False}

    def open_folder(self, path):
        """调用系统资源管理器打开特定文件夹"""
        import os
        import platform
        import subprocess

        # 确保存储的路径分隔符是环境兼容的
        path = os.path.normpath(path)
        print(f"Opening folder: {path}")

        if not os.path.exists(path):
            try:
                os.makedirs(path, exist_ok=True)
            except Exception as e:
                 return {"success": False, "message": f"无法创建目录: {str(e)}"}

        try:
            if platform.system() == "Windows":
                # Windows：推荐采用 subprocess.Popen 或者 os.startfile 这里采用 explorer 确保前台弹出
                subprocess.Popen(f'explorer "{path}"')
            elif platform.system() == "Darwin":
                # macOS
                subprocess.Popen(["open", path])
            else:
                # Linux
                subprocess.Popen(["xdg-open", path])
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def retry_all_errors(self):
        """重新处理所有失败的发票任务（占位实现）"""
        print("Retrying all failed invoices...")
        return {"success": True, "message": "已为您触发全部重试任务"}

    def retry_single_invoice(self, name):
        """重新处理特定的一张发票（占位实现）"""
        print(f"Retrying single invoice: {name}")
        return {"success": True, "message": f"正在重试发票: {name}"}

    def view_invoice(self, path):
        """在系统默认查看器中打开单张发票或图片"""
        print(f"Viewing invoice: {path}")
        import os
        import platform
        import subprocess
        
        path = os.path.normpath(path)
        if not os.path.exists(path):
            return {"success": False, "message": "文件不存在"}
            
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
            return {"success": True, "message": "已在外部程序中打开文件"}
        except Exception as e:
            return {"success": False, "message": f"打开失败: {str(e)}"}

    def export_run_summary(self, export_path=""):
        try:
            from openpyxl import Workbook
        except ImportError:
            return {"success": False, "message": "缺少 openpyxl 依赖，无法导出 Excel"}

        target_dir = export_path or self._effective_save_path or self._requested_save_path or self.get_default_save_path()
        os.makedirs(target_dir, exist_ok=True)

        def _parse_amount(value):
            text = str(value or "").strip()
            if not text:
                return 0.0
            cleaned = re.sub(r"[^\d.\-]", "", text)
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except Exception:
                return 0.0

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "分类汇总"
        summary_sheet.append(["发票分类", "数量", "金额合计"])

        category_summary = {}
        for item in self.processed_invoices:
            category = str(item.get("category") or "未分类")
            bucket = category_summary.setdefault(category, {"count": 0, "amount": 0.0})
            bucket["count"] += 1
            bucket["amount"] += _parse_amount(item.get("amount", ""))

        if category_summary:
            for category in sorted(category_summary.keys()):
                bucket = category_summary[category]
                summary_sheet.append([category, bucket["count"], round(bucket["amount"], 2)])
        else:
            summary_sheet.append(["暂无成功记录", 0, 0.0])

        success_sheet = workbook.create_sheet("成功明细")
        success_sheet.append(["日期", "金额", "销售方", "分类", "文件路径"])
        for item in self.processed_invoices:
            success_sheet.append([
                item.get("date", ""),
                item.get("amount", ""),
                item.get("merchant", ""),
                item.get("category", ""),
                item.get("path", ""),
            ])

        error_sheet = workbook.create_sheet("异常记录")
        error_sheet.append(["分组", "状态", "原因", "日期", "金额", "销售方", "文件名", "文件路径"])
        for group in self._group_error_invoices():
            for item in group.get("items", []):
                error_sheet.append([
                    group.get("label", ""),
                    item.get("status", ""),
                    item.get("reason", ""),
                    item.get("date", ""),
                    item.get("amount", ""),
                    item.get("merchant", ""),
                    item.get("name", ""),
                    item.get("path", ""),
                ])

        export_file = os.path.join(target_dir, f"结果明细_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        workbook.save(export_file)
        self._last_export_path = export_file
        return {"success": True, "message": "结果明细已导出", "path": export_file}

if __name__ == "__main__":
    # 提供 CLI 测试入口
    print(">>> [测试入口] 从终端直接启动 app_api.py")
    app = InvoiceAppAPI()
    try:
        import time
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print(">>> 退出程序。")
